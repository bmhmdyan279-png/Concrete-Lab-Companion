#!/usr/bin/env python3
"""
Concrete Lab Companion — GENERATION v2.0.0
==========================================
Professional Excel workbook for concrete testing.
Fully data‑driven, maintainable, and standards‑compliant.
"""

import argparse
import hashlib
import logging
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Protection, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ─── Logging ────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ─── Constants ──────────────────────────────────────────────────────────────
VERSION = "2.0.0"
BUILD_DATE = datetime.now().strftime("%Y-%m-%d %H:%M")
PASSWORD = os.getenv("WORKBOOK_PASSWORD", "ConcreteLab2026!")

# Standard specifications (ASTM / ISIRI)
STANDARDS = {
    "C29": {"name": "ASTM C29/C29M", "edition": "2023", "title": "Bulk Density of Aggregates"},
    "C39": {"name": "ASTM C39/C39M", "edition": "2026", "title": "Compressive Strength of Concrete"},
    "C78": {"name": "ASTM C78/C78M", "edition": "2022", "title": "Flexural Strength (Third-Point)"},
    "C127": {"name": "ASTM C127", "edition": "2023", "title": "Density of Coarse Aggregate"},
    "C128": {"name": "ASTM C128", "edition": "2022", "title": "Density of Fine Aggregate"},
    "C136": {"name": "ASTM C136/C136M", "edition": "2024", "title": "Sieve Analysis"},
    "C138": {"name": "ASTM C138/C138M", "edition": "2024", "title": "Density of Fresh Concrete"},
    "C143": {"name": "ASTM C143/C143M", "edition": "2024", "title": "Slump of Hydraulic Cement Concrete"},
    "C187": {"name": "ASTM C187", "edition": "2023", "title": "Normal Consistency of Hydraulic Cement"},
    "C191": {"name": "ASTM C191", "edition": "2024", "title": "Setting Time by Vicat Needle"},
    "C232": {"name": "ASTM C232/C232M", "edition": "2023", "title": "Bleeding of Concrete"},
    "C293": {"name": "ASTM C293/C293M", "edition": "2023", "title": "Flexural Strength (Center-Point)"},
    "C496": {"name": "ASTM C496/C496M", "edition": "2023", "title": "Splitting Tensile Strength"},
    "C566": {"name": "ASTM C566", "edition": "2023", "title": "Moisture Content of Aggregates"},
    "C597": {"name": "ASTM C597", "edition": "2023", "title": "Pulse Velocity Through Concrete"},
    "C805": {"name": "ASTM C805/C805M", "edition": "2025", "title": "Rebound Hammer"},
    "D2419": {"name": "ASTM D2419", "edition": "2022", "title": "Sand Equivalent"},
    "D4791": {"name": "ASTM D4791", "edition": "2023", "title": "Flat Particles, Elongated Particles"},
    "EN196-1": {"name": "EN 196-1", "edition": "2023", "title": "Mortar Strength"},
    "ISIRI302": {"name": "ISIRI 302", "edition": "2020", "title": "Aggregate Grading Limits"},
}

# ─── Style System ──────────────────────────────────────────────────────────
class StyleManager:
    """Centralised style definitions."""
    COLORS = {
        "input_fill": "FFF2CC", "input_border": "D9D9D9", "calc_fill": "F2F2F2",
        "pass_fill": "C6EFCE", "pass_font": "006100", "warn_fill": "FCE4D6",
        "warn_font": "C00000", "fail_fill": "FFC7CE", "fail_font": "9C0006",
        "header_fill": "1F4E79", "header_font": "FFFFFF", "nav_fill": "D6E4F0",
    }
    FONT_BODY = "Tahoma"
    FONT_NUM = "Calibri"

    def __init__(self):
        self.styles = {}
        self._build_styles()

    def _build_styles(self):
        c = self.COLORS
        self.styles.update({
            "header": {
                "fill": PatternFill("solid", fgColor=c["header_fill"]),
                "font": Font(name=self.FONT_BODY, color=c["header_font"], bold=True, size=11),
            },
            "input": {
                "fill": PatternFill("solid", fgColor=c["input_fill"]),
                "border": Border(
                    left=Side("thin", c["input_border"]), right=Side("thin", c["input_border"]),
                    top=Side("thin", c["input_border"]), bottom=Side("thin", c["input_border"])
                ),
            },
            "calc": {
                "fill": PatternFill("solid", fgColor=c["calc_fill"]),
            },
            "pass": {
                "fill": PatternFill("solid", fgColor=c["pass_fill"]),
                "font": Font(name=self.FONT_BODY, color=c["pass_font"], bold=True),
            },
            "warn": {
                "fill": PatternFill("solid", fgColor=c["warn_fill"]),
                "font": Font(name=self.FONT_BODY, color=c["warn_font"], bold=True),
            },
            "fail": {
                "fill": PatternFill("solid", fgColor=c["fail_fill"]),
                "font": Font(name=self.FONT_BODY, color=c["fail_font"], bold=True),
            },
            "title": {
                "font": Font(name=self.FONT_BODY, bold=True, size=14, color=c["header_fill"]),
            },
            "subtitle": {
                "font": Font(name=self.FONT_BODY, size=9, italic=True, color="666666"),
            },
            "nav": {
                "fill": PatternFill("solid", fgColor=c["nav_fill"]),
                "font": Font(name=self.FONT_BODY, size=10, color=c["header_fill"]),
            },
            "normal": {
                "font": Font(name=self.FONT_BODY, size=11),
            },
            "num": {
                "font": Font(name=self.FONT_NUM, size=11),
            },
            "thin_border": Border(
                left=Side("thin", "BFBFBF"), right=Side("thin", "BFBFBF"),
                top=Side("thin", "BFBFBF"), bottom=Side("thin", "BFBFBF")
            ),
            "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "right": Alignment(horizontal="right", vertical="center", wrap_text=True),
        })

    def get(self, key: str) -> Dict:
        return self.styles.get(key, {})

    def apply(self, cell, style_key: str, extra: Dict = None):
        style = self.get(style_key)
        if not style:
            return
        if "font" in style:
            cell.font = style["font"]
        if "fill" in style:
            cell.fill = style["fill"]
        if "border" in style:
            cell.border = style["border"]
        if "alignment" in style:
            cell.alignment = style["alignment"]
        if extra:
            for k, v in extra.items():
                setattr(cell, k, v)


STYLES = StyleManager()

# ─── Test Specification ─────────────────────────────────────────────────────
class TestStatus(Enum):
    NOT_READY = "—"
    PASS = "✅"
    WARN = "⚠️"
    FAIL = "❌"


@dataclass
class InputField:
    key: str
    label: str
    row: int
    col: int = 3
    unit: str = ""
    validation: Dict[str, Any] = field(default_factory=dict)
    tooltip: str = ""

@dataclass
class OutputField:
    key: str
    label: str
    row: int
    col: int
    formula: str
    unit: str = ""
    num_fmt: str = "0.00"
    style: str = "calc"

@dataclass
class Check:
    label: str
    formula: str
    row: int
    col: int = 3
    style_pass: str = "pass"
    style_fail: str = "fail"

@dataclass
class TestSpec:
    id: str
    title: str
    standard: str
    tab_color: str
    sheet_name: str
    inputs: List[InputField] = field(default_factory=list)
    outputs: List[OutputField] = field(default_factory=list)
    checks: List[Check] = field(default_factory=list)
    chart: Optional[Dict] = None


# ─── Test Registry ─────────────────────────────────────────────────────────
TEST_REGISTRY: Dict[str, TestSpec] = {}

def register_test(spec: TestSpec) -> TestSpec:
    TEST_REGISTRY[spec.id] = spec
    return spec

# ─── Define all tests ──────────────────────────────────────────────────────
def define_tests():
    # 1-1 Sieve Analysis
    register_test(TestSpec(
        id="1-1",
        title="دانه‌بندی سنگدانه",
        standard=STANDARDS["C136"]["name"],
        tab_color="FF6F00",
        sheet_name="02_آزمایش_1-1",
        inputs=[
            InputField("initial_mass", "جرم اولیه نمونه خشک (g):", 5, validation={"type":"decimal","min":0,"max":100000}),
        ],
        outputs=[
            OutputField("fm", "مدول نرمی (FM):", 19, 3,
                        '=IF(OR($C$5=0,C8=""),"—",ROUND(SUMPRODUCT((G8:G16=TRUE)*E8:E16)/100,2))',
                        num_fmt="0.00", style="pass"),
        ],
        checks=[
            Check("کنترل جرم", '=IF($C$5=0,"—",IF(ABS($C$5-SUM(C8:C16))/$C$5>0.003,"❌ اختلاف >0.3%","✅"))', 18),
        ],
        chart={
            "type": "scatter",
            "x_col": 1, "y_col": 2,
            "data_start": 22, "data_end": 29,
            "title": "منحنی دانه‌بندی",
        }
    ))

    # 1-2 Moisture
    register_test(TestSpec(
        id="1-2",
        title="رطوبت سنگدانه",
        standard=STANDARDS["C566"]["name"],
        tab_color="FF6F00",
        sheet_name="03_آزمایش_1-2",
        inputs=[
            InputField("W1", "جرم نمونه تر (g):", 5, validation={"type":"decimal","min":0,"max":100000}),
            InputField("W2", "جرم نمونه خشک (g):", 6, validation={"type":"decimal","min":0,"max":100000}),
        ],
        outputs=[
            OutputField("moisture", "درصد رطوبت (پایه خشک):", 8, 3,
                        '=IF(OR(C5="",C6="",C6=0),"—",ROUND((C5-C6)/C6*100,2))',
                        num_fmt="0.00"),
        ],
        checks=[]
    ))

    # 1-3 Coarse aggregate density
    register_test(TestSpec(
        id="1-3",
        title="چگالی سنگدانه درشت",
        standard=STANDARDS["C127"]["name"],
        tab_color="FF6F00",
        sheet_name="04_آزمایش_1-3",
        inputs=[
            InputField("A", "جرم خشک (g):", 5, validation={"type":"decimal","min":0,"max":100000}),
            InputField("B", "جرم SSD (g):", 6, validation={"type":"decimal","min":0,"max":100000}),
            InputField("C", "جرم در آب (g):", 7, validation={"type":"decimal","min":0,"max":100000}),
        ],
        outputs=[
            OutputField("OD", "چگالی خشک (OD):", 9, 3,
                        '=IF(OR(C5="",C6="",C7="",C6=C7),"—",ROUND(C5/(C6-C7),3))',
                        num_fmt="0.000"),
            OutputField("SSD", "چگالی SSD:", 10, 3,
                        '=IF(OR(C5="",C6="",C7="",C6=C7),"—",ROUND(C6/(C6-C7),3))',
                        num_fmt="0.000"),
            OutputField("App", "چگالی ظاهری:", 11, 3,
                        '=IF(OR(C5="",C6="",C7="",C5=C7),"—",ROUND(C5/(C5-C7),3))',
                        num_fmt="0.000"),
            OutputField("Abs", "جذب آب (%):", 12, 3,
                        '=IF(OR(C5="",C6="",C5=0),"—",ROUND((C6-C5)/C5*100,2))',
                        num_fmt="0.00"),
        ],
        checks=[
            Check("بررسی فیزیکی (SSD ≥ OD)",
                  '=IF(OR(C9="—",C10="—"),"—",IF(C10<C9-0.01,"❌ غیرفیزیکی","✅"))', 14),
            Check("بررسی فیزیکی (App ≥ SSD)",
                  '=IF(OR(C10="—",C11="—"),"—",IF(C11<C10-0.01,"❌ غیرفیزیکی","✅"))', 15),
        ]
    ))

    # 1-4 Fine aggregate density
    register_test(TestSpec(
        id="1-4",
        title="چگالی سنگدانه ریز",
        standard=STANDARDS["C128"]["name"],
        tab_color="FF6F00",
        sheet_name="05_آزمایش_1-4",
        inputs=[
            InputField("A", "جرم خشک (g):", 7, validation={"type":"decimal","min":0,"max":100000}),
            InputField("S", "جرم SSD (g):", 8, validation={"type":"decimal","min":0,"max":100000}),
            InputField("B", "جرم ظرف+آب (g):", 9, validation={"type":"decimal","min":0,"max":100000}),
            InputField("C", "جرم ظرف+آب+نمونه (g):", 10, validation={"type":"decimal","min":0,"max":100000}),
        ],
        outputs=[
            OutputField("OD", "چگالی خشک (OD):", 12, 3,
                        '=IF(OR(C7="",C9="",C10="",C9+C8-C10=0),"—",ROUND(C7/(C9+C8-C10),3))',
                        num_fmt="0.000"),
            OutputField("SSD", "چگالی SSD:", 13, 3,
                        '=IF(OR(C8="",C9="",C10="",C9+C8-C10=0),"—",ROUND(C8/(C9+C8-C10),3))',
                        num_fmt="0.000"),
            OutputField("App", "چگالی ظاهری:", 14, 3,
                        '=IF(OR(C7="",C9="",C10="",C9+C7-C10=0),"—",ROUND(C7/(C9+C7-C10),3))',
                        num_fmt="0.000"),
            OutputField("Abs", "جذب آب (%):", 15, 3,
                        '=IF(OR(C7="",C8="",C7=0),"—",ROUND((C8-C7)/C7*100,2))',
                        num_fmt="0.00"),
        ],
        checks=[
            Check("بررسی فیزیکی (SSD ≥ OD)",
                  '=IF(OR(C12="—",C13="—"),"—",IF(C13<C12-0.01,"❌ غیرفیزیکی","✅"))', 17),
            Check("بررسی فیزیکی (App ≥ SSD)",
                  '=IF(OR(C13="—",C14="—"),"—",IF(C14<C13-0.01,"❌ غیرفیزیکی","✅"))', 18),
        ]
    ))

    # 1-5 Unit weight of aggregate (ASTM C29)
    register_test(TestSpec(
        id="1-5",
        title="وزن واحد حجمی سنگدانه",
        standard=STANDARDS["C29"]["name"],
        tab_color="FF6F00",
        sheet_name="06_آزمایش_1-5",
        inputs=[
            InputField("T", "جرم ظرف خالی (g):", 5, validation={"type":"decimal","min":0,"max":100000}),
            InputField("G", "جرم ظرف+سنگدانه (g):", 6, validation={"type":"decimal","min":0,"max":100000}),
            InputField("V", "حجم ظرف (cm³):", 7, validation={"type":"decimal","min":0,"max":100000}),
            InputField("S", "چگالی (بی‌بُعد):", 8, validation={"type":"decimal","min":2,"max":3.5}),
        ],
        outputs=[
            OutputField("UW", "وزن واحد حجمی (kg/m³):", 10, 3,
                        '=IF(OR(C5="",C6="",C7="",C7=0),"—",ROUND((C6-C5)/C7*1000,0))',
                        num_fmt="#,##0"),
            OutputField("Voids", "فضای خالی (%):", 11, 3,
                        '=IF(OR(C8="",C8=0,C9="—"),"—",ROUND((C8*1000-C9)/(C8*1000)*100,1))',
                        num_fmt="0.0"),
        ],
        checks=[]
    ))

    # 1-6 Sand Equivalent (ASTM D2419) — corrected formula
    register_test(TestSpec(
        id="1-6",
        title="معادل ماسه",
        standard=STANDARDS["D2419"]["name"],
        tab_color="FF6F00",
        sheet_name="07_آزمایش_1-6",
        inputs=[
            InputField("Sand", "خوانش ماسه (mm):", 5, validation={"type":"decimal","min":0,"max":500}),
            InputField("Clay", "خوانش رس (mm):", 6, validation={"type":"decimal","min":0,"max":500}),
        ],
        outputs=[
            OutputField("SE", "معادل ماسه SE (%):", 8, 3,
                        '=IF(OR(C5="",C6="",C6=0),"—",ROUNDUP(C5/(C5+C6)*100,0))',
                        num_fmt="0"),
        ],
        checks=[]
    ))

    # 1-7 Shape indices
    register_test(TestSpec(
        id="1-7",
        title="شاخص‌های شکل سنگدانه",
        standard=STANDARDS["D4791"]["name"],
        tab_color="FF6F00",
        sheet_name="08_آزمایش_1-7",
        inputs=[
            InputField("Wtot", "جرم کل (g):", 5, validation={"type":"decimal","min":0,"max":100000}),
            InputField("Wlong", "جرم دراز (g):", 6, validation={"type":"decimal","min":0,"max":100000}),
            InputField("Wflat", "جرم پهن (g):", 7, validation={"type":"decimal","min":0,"max":100000}),
            InputField("Wboth", "جرم هر دو (g):", 8, validation={"type":"decimal","min":0,"max":100000}),
        ],
        outputs=[
            OutputField("LI", "شاخص درازگی (%):", 10, 3,
                        '=IF(OR(C5="",C5=0,C6=""),"—",ROUND(C6/C5*100,1))', num_fmt="0.0"),
            OutputField("FI", "شاخص پهنی (%):", 11, 3,
                        '=IF(OR(C5="",C5=0,C7=""),"—",ROUND(C7/C5*100,1))', num_fmt="0.0"),
            OutputField("BI", "شاخص هر دو (%):", 12, 3,
                        '=IF(OR(C5="",C5=0,C8=""),"—",ROUND(C8/C5*100,1))', num_fmt="0.0"),
        ],
        checks=[]
    ))

    # 1-8 Absorption
    register_test(TestSpec(
        id="1-8",
        title="جذب آب سنگدانه",
        standard=STANDARDS["C127"]["name"],
        tab_color="FF6F00",
        sheet_name="09_آزمایش_1-8",
        inputs=[
            InputField("W1", "جرم SSD (g):", 5, validation={"type":"decimal","min":0,"max":100000}),
            InputField("W2", "جرم خشک (g):", 6, validation={"type":"decimal","min":0,"max":100000}),
        ],
        outputs=[
            OutputField("Abs", "جذب آب (%):", 8, 3,
                        '=IF(OR(C5="",C6="",C6=0),"—",ROUND((C5-C6)/C6*100,2))',
                        num_fmt="0.00"),
        ],
        checks=[]
    ))

    # 2-1 Fresh concrete density (ASTM C138)
    register_test(TestSpec(
        id="2-1",
        title="چگالی بتن تازه",
        standard=STANDARDS["C138"]["name"],
        tab_color="2196F3",
        sheet_name="10_آزمایش_2-1",
        inputs=[
            InputField("Ma", "جرم ظرف خالی (g):", 5, validation={"type":"decimal","min":0,"max":100000}),
            InputField("Mt", "جرم ظرف+بتن (g):", 6, validation={"type":"decimal","min":0,"max":100000}),
            InputField("V", "حجم ظرف (cm³):", 7, validation={"type":"decimal","min":0,"max":100000}),
        ],
        outputs=[
            OutputField("Density", "چگالی (kg/m³):", 9, 3,
                        '=IF(OR(C5="",C6="",C7="",C7=0),"—",ROUND((C6-C5)/C7*1000,3))',
                        num_fmt="0.000"),
        ],
        checks=[]
    ))

    # 2-2 Normal Consistency (ASTM C187)
    register_test(TestSpec(
        id="2-2",
        title="قوام نرمال سیمان (ویکات)",
        standard=STANDARDS["C187"]["name"],
        tab_color="2196F3",
        sheet_name="11_آزمایش_2-2",
        inputs=[
            InputField("Cement", "سیمان (g):", 5, validation={"type":"decimal","min":0,"max":10000}),
            InputField("Water", "آب (g):", 6, validation={"type":"decimal","min":0,"max":10000}),
            InputField("Penetration", "نفوذ اولیه (mm):", 7, validation={"type":"decimal","min":0,"max":50}),
        ],
        outputs=[
            OutputField("WC", "نسبت آب به سیمان (%):", 9, 3,
                        '=IF(OR(C5="",C6="",C5=0),"—",ROUND(C6/C5*100,1))', num_fmt="0.0"),
        ],
        checks=[
            Check("وضعیت نفوذ",
                  '=IF(C7="","",IF(ABS(C7-10)>1,"⚠️ تکرار با آب جدید","✅"))', 10),
        ]
    ))

    # 2-3 Setting Time (ASTM C191)
    register_test(TestSpec(
        id="2-3",
        title="زمان گیرش سیمان",
        standard=STANDARDS["C191"]["name"],
        tab_color="2196F3",
        sheet_name="12_آزمایش_2-3",
        inputs=[
            InputField("E", "زمان اولیه (min):", 5, validation={"type":"decimal","min":0,"max":10000}),
            InputField("H", "زمان ثانویه (min):", 6, validation={"type":"decimal","min":0,"max":10000}),
            InputField("C", "نفوذ ثانویه (mm):", 7, validation={"type":"decimal","min":0,"max":50}),
            InputField("D", "نفوذ اولیه (mm):", 8, validation={"type":"decimal","min":0,"max":50}),
            InputField("Final", "زمان گیرش نهایی (min):", 9, validation={"type":"decimal","min":0,"max":10000}),
        ],
        outputs=[
            OutputField("Initial", "زمان گیرش اولیه (min):", 11, 3,
                        '=IF(OR(C5="",C6="",C7="",C8="",C7=C8),"—",ROUND(C5+(C6-C5)*(25-C8)/(C7-C8),0))',
                        num_fmt="0"),
            OutputField("FinalM", "زمان گیرش نهایی (min):", 12, 3,
                        '=IF(C9="","",MROUND(C9,5))', num_fmt="0"),
        ],
        checks=[]
    ))

    # 2-4 Mortar strength (EN 196-1)
    register_test(TestSpec(
        id="2-4",
        title="مقاومت ملات سیمان",
        standard=STANDARDS["EN196-1"]["name"],
        tab_color="2196F3",
        sheet_name="13_آزمایش_2-4",
        inputs=[
            InputField("F1", "بار خمشی ۱ (kgf):", 6, col=1, validation={"type":"decimal","min":0,"max":10000}),
            InputField("F2", "بار خمشی ۲ (kgf):", 6, col=2, validation={"type":"decimal","min":0,"max":10000}),
            InputField("F3", "بار خمشی ۳ (kgf):", 6, col=3, validation={"type":"decimal","min":0,"max":10000}),
            InputField("C1", "بار فشاری ۱ (kgf):", 10, col=1, validation={"type":"decimal","min":0,"max":50000}),
            InputField("C2", "بار فشاری ۲ (kgf):", 10, col=2, validation={"type":"decimal","min":0,"max":50000}),
            InputField("C3", "بار فشاری ۳ (kgf):", 10, col=3, validation={"type":"decimal","min":0,"max":50000}),
            InputField("C4", "بار فشاری ۴ (kgf):", 10, col=4, validation={"type":"decimal","min":0,"max":50000}),
            InputField("C5", "بار فشاری ۵ (kgf):", 10, col=5, validation={"type":"decimal","min":0,"max":50000}),
            InputField("C6", "بار فشاری ۶ (kgf):", 10, col=6, validation={"type":"decimal","min":0,"max":50000}),
        ],
        outputs=[
            OutputField("Flex", "مقاومت خمشی (MPa):", 8, 4,
                        '=IF(COUNT(A6:C6)<3,"—",ROUND(AVERAGE(1.5*A6*9.80665*100/40^3,1.5*B6*9.80665*100/40^3,1.5*C6*9.80665*100/40^3),1))',
                        num_fmt="0.0"),
            OutputField("Comp", "مقاومت فشاری (MPa):", 12, 4,
                        '=IF(COUNT(A10:F10)<6,"—",ROUND(AVERAGE(A10*9.80665/1600,B10*9.80665/1600,C10*9.80665/1600,D10*9.80665/1600,E10*9.80665/1600,F10*9.80665/1600),1))',
                        num_fmt="0.0"),
        ],
        checks=[]
    ))

    # 3-1 Slump (height correction: 304.8 mm)
    register_test(TestSpec(
        id="3-1",
        title="اسلامپ بتن",
        standard=STANDARDS["C143"]["name"],
        tab_color="4CAF50",
        sheet_name="14_آزمایش_3-1",
        inputs=[
            InputField("H", "ارتفاع پس از برداشتن (mm):", 5, validation={"type":"decimal","min":0,"max":300}),
            InputField("Type", "نوع ریزش:", 6, validation={"type":"list","formula":"برشی;دو طرفه;ریزش کامل"}),
        ],
        outputs=[
            OutputField("Slump", "اسلامپ (mm):", 8, 3,
                        '=IF(C5="","",MROUND(304.8-C5,5))', num_fmt="0"),
        ],
        checks=[
            Check("وضعیت", '=IF(C6="","",IF(C6="ریزش کامل","⚠️ تکرار آزمایش","✅"))', 9),
        ]
    ))

    # 3-2 Bleeding (approximate height-based)
    register_test(TestSpec(
        id="3-2",
        title="آب‌انداختگی بتن",
        standard=STANDARDS["C232"]["name"],
        tab_color="4CAF50",
        sheet_name="15_آزمایش_3-2",
        inputs=[
            InputField("h1", "ارتفاع اولیه (mm):", 5, validation={"type":"decimal","min":0,"max":1000}),
            InputField("h2", "ارتفاع نهایی (mm):", 6, validation={"type":"decimal","min":0,"max":1000}),
            InputField("G", "جذب سنگدانه (%):", 7, validation={"type":"decimal","min":0,"max":1000}),
        ],
        outputs=[
            OutputField("Bleed", "آب‌انداختگی ظاهری (%):", 9, 3,
                        '=IF(OR(C5="",C6="",C5=0),"—",ROUND((C5-C6)/C5*100,1))', num_fmt="0.0"),
            OutputField("BleedReal", "آب‌انداختگی واقعی (%):", 10, 3,
                        '=IF(OR(C9="—",C7=""),"—",ROUND(C9-C7,1))', num_fmt="0.0"),
        ],
        checks=[]
    ))

    # 3-3 Unit weight of concrete (ASTM C138)
    register_test(TestSpec(
        id="3-3",
        title="وزن واحد حجمی بتن",
        standard=STANDARDS["C138"]["name"],
        tab_color="4CAF50",
        sheet_name="16_آزمایش_3-3",
        inputs=[
            InputField("m1", "جرم ظرف خالی (g):", 5, validation={"type":"decimal","min":0,"max":100000}),
            InputField("m2", "جرم ظرف+بتن (g):", 6, validation={"type":"decimal","min":0,"max":100000}),
            InputField("V", "حجم ظرف (cm³):", 7, validation={"type":"decimal","min":0,"max":100000}),
            InputField("D_theo", "چگالی نظری (kg/m³):", 8, validation={"type":"decimal","min":0,"max":10000}),
        ],
        outputs=[
            OutputField("UW", "وزن واحد حجمی (kg/m³):", 10, 3,
                        '=IF(OR(C5="",C6="",C7="",C7=0),"—",ROUND((C6-C5)/C7*1000,0))',
                        num_fmt="#,##0"),
            OutputField("Dev", "اختلاف با نظری (%):", 11, 3,
                        '=IF(OR(C10="—",C8="",C8=0),"—",ROUND(ABS(C10-C8)/C8*100,1))',
                        num_fmt="0.0"),
        ],
        checks=[
            Check("وضعیت", '=IF(C11="—","",IF(C11>2,"⚠️ اختلاف >2%","✅"))', 12),
        ]
    ))

    # 4-1 Compressive strength (ASTM C39 — only cylinders)
    register_test(TestSpec(
        id="4-1",
        title="مقاومت فشاری بتن (استوانه)",
        standard=STANDARDS["C39"]["name"],
        tab_color="F44336",
        sheet_name="17_آزمایش_4-1",
        inputs=[
            InputField("D", "قطر (mm):", 6, validation={"type":"decimal","min":0,"max":500}),
            InputField("P", "بار (kN):", 7, validation={"type":"decimal","min":0,"max":10000}),
            InputField("Pattern", "الگوی شکست:", 8, validation={"type":"list","formula":"نوع ۱;نوع ۲;نوع ۳;نوع ۴;نوع ۵;نوع ۶"}),
        ],
        outputs=[
            OutputField("Area", "مساحت (mm²):", 10, 3,
                        '=IF(C6="","",PI()/4*C6^2)', num_fmt="0.0"),
            OutputField("Strength", "مقاومت فشاری (MPa):", 11, 3,
                        '=IF(OR(C7="",C10="—",C10="",C10=0),"—",MROUND(C7*1000/C10,0.1))',
                        num_fmt="0.1", style="pass"),
        ],
        checks=[
            Check("الگوی شکست معتبر",
                  '=IF(C8="","",IF(OR(C8="نوع ۱",C8="نوع ۲",C8="نوع ۳"),"✅ معتبر","⚠️ نامعتبر"))', 12),
        ]
    ))

    # 4-2 Splitting tensile (ASTM C496)
    register_test(TestSpec(
        id="4-2",
        title="مقاومت کششی (برزیلی)",
        standard=STANDARDS["C496"]["name"],
        tab_color="F44336",
        sheet_name="18_آزمایش_4-2",
        inputs=[
            InputField("d", "قطر (mm):", 5, validation={"type":"decimal","min":0,"max":1000}),
            InputField("L", "طول (mm):", 6, validation={"type":"decimal","min":0,"max":1000}),
            InputField("P", "بار (N):", 7, validation={"type":"decimal","min":0,"max":1000000}),
        ],
        outputs=[
            OutputField("Fct", "مقاومت کششی (MPa):", 9, 3,
                        '=IF(OR(C5="",C6="",C7="",C5=0,C6=0),"—",ROUND(2*C7/(PI()*C5*C6),2))',
                        num_fmt="0.00"),
        ],
        checks=[
            Check("بازه منطقی (۲-۸ MPa)",
                  '=IF(C9="—","",IF(OR(C9<2,C9>8),"⚠️ خارج از بازه","✅"))', 10),
        ]
    ))

    # 4-3 Flexural strength (ASTM C78 third-point + C293 center-point)
    register_test(TestSpec(
        id="4-3",
        title="مقاومت خمشی",
        standard=STANDARDS["C78"]["name"] + " / " + STANDARDS["C293"]["name"],
        tab_color="F44336",
        sheet_name="19_آزمایش_4-3",
        inputs=[
            InputField("b", "عرض (mm):", 5, validation={"type":"decimal","min":0,"max":1000}),
            InputField("d", "ارتفاع (mm):", 6, validation={"type":"decimal","min":0,"max":1000}),
            InputField("L", "دهانه (mm):", 7, validation={"type":"decimal","min":0,"max":1000}),
            InputField("P", "بار (N):", 8, validation={"type":"decimal","min":0,"max":1000000}),
            InputField("Method", "روش بارگذاری:", 10, validation={"type":"list","formula":"یک‌سوم میانه;مرکزی"}),
            InputField("Crack", "محل ترک:", 11, validation={"type":"list","formula":"داخل محدوده;خارج محدوده"}),
        ],
        outputs=[
            OutputField("Flex", "مقاومت خمشی (MPa):", 13, 3,
                        '=IF(OR(C5="",C6="",C7="",C8="",C11="خارج محدوده"),IF(C11="خارج محدوده","⚠️ تکرار","—"),IF(C10="مرکزی",ROUND(3*C8*C7/(2*C5*C6^2),2),ROUND(C8*C7/(C5*C6^2),2)))',
                        num_fmt="0.00"),
        ],
        checks=[]
    ))

    # 4-4 UPV (ASTM C597) — with cautious interpretation
    register_test(TestSpec(
        id="4-4",
        title="سرعت پالس اولتراسونیک",
        standard=STANDARDS["C597"]["name"],
        tab_color="F44336",
        sheet_name="20_آزمایش_4-4",
        inputs=[
            InputField("L", "طول مسیر (m):", 5, validation={"type":"decimal","min":0,"max":10}),
            InputField("T1", "زمان ۱ (µs):", 6, col=3, validation={"type":"decimal","min":0,"max":100000}),
            InputField("T2", "زمان ۲ (µs):", 6, col=4, validation={"type":"decimal","min":0,"max":100000}),
            InputField("T3", "زمان ۳ (µs):", 6, col=5, validation={"type":"decimal","min":0,"max":100000}),
        ],
        outputs=[
            OutputField("Velocity", "سرعت (km/s):", 8, 3,
                        '=IF(OR(C5="",C6="",C6=0),"—",ROUND(C5*1000/AVERAGE(C6:E6),2))',
                        num_fmt="0.00"),
            OutputField("Stdev", "انحراف معیار:", 9, 3,
                        '=IF(C6="","",IFERROR(ROUND(STDEV(C6:E6),2),"—"))', num_fmt="0.00"),
        ],
        checks=[
            Check("طبقه‌بندی کیفی (مرجع آموزشی)",
                  '=IF(C8="—","",IF(C8>=4.5,"عالی",IF(C8>=3.5,"خوب",IF(C8>=3,"متوسط","ضعیف"))))', 10),
        ]
    ))

    # 4-5 Schmidt rebound (ASTM C805) — with angle correction
    register_test(TestSpec(
        id="4-5",
        title="چکش اشمیت",
        standard=STANDARDS["C805"]["name"],
        tab_color="F44336",
        sheet_name="21_آزمایش_4-5",
        inputs=[
            InputField("R1", "R1", 6, col=1, validation={"type":"decimal","min":0,"max":100}),
            InputField("R2", "R2", 6, col=2, validation={"type":"decimal","min":0,"max":100}),
            InputField("R3", "R3", 6, col=3, validation={"type":"decimal","min":0,"max":100}),
            InputField("R4", "R4", 6, col=4, validation={"type":"decimal","min":0,"max":100}),
            InputField("R5", "R5", 6, col=5, validation={"type":"decimal","min":0,"max":100}),
            InputField("R6", "R6", 6, col=6, validation={"type":"decimal","min":0,"max":100}),
            InputField("R7", "R7", 6, col=7, validation={"type":"decimal","min":0,"max":100}),
            InputField("R8", "R8", 6, col=8, validation={"type":"decimal","min":0,"max":100}),
            InputField("R9", "R9", 7, col=1, validation={"type":"decimal","min":0,"max":100}),
            InputField("R10", "R10", 7, col=2, validation={"type":"decimal","min":0,"max":100}),
            InputField("R11", "R11", 7, col=3, validation={"type":"decimal","min":0,"max":100}),
            InputField("R12", "R12", 7, col=4, validation={"type":"decimal","min":0,"max":100}),
            InputField("R13", "R13", 7, col=5, validation={"type":"decimal","min":0,"max":100}),
            InputField("R14", "R14", 7, col=6, validation={"type":"decimal","min":0,"max":100}),
            InputField("R15", "R15", 7, col=7, validation={"type":"decimal","min":0,"max":100}),
            InputField("R16", "R16", 7, col=8, validation={"type":"decimal","min":0,"max":100}),
            InputField("Surface", "سطح:", 9, validation={"type":"list","formula":"خشک;مرطوب"}),
            InputField("Angle", "زاویه ضربه:", 10, validation={"type":"list","formula":"عمودی به پایین-90;عمودی به بالا+90;افقی 0"}),
            InputField("Temp", "دما (°C):", 11, validation={"type":"decimal","min":-10,"max":60}),
        ],
        outputs=[
            OutputField("Avg", "میانگین کل:", 13, 3,
                        '=IF(COUNT(A6:H7)=0,"",ROUND(AVERAGE(A6:H7),1))', num_fmt="0.0"),
            OutputField("Nvalid", "تعداد معتبر:", 14, 3,
                        '=IF(C13="","",COUNTIFS(A6:H7, ">=" & C13-6, A6:H7, "<=" & C13+6))', num_fmt="0"),
            OutputField("Rm", "میانگین معتبرها:", 15, 3,
                        '=IF(OR(C13="",C14=0),"—",ROUND(SUMPRODUCT((ABS(IF(ISNUMBER(A6:H7),A6:H7,999)-C13)<=6)*IF(ISNUMBER(A6:H7),A6:H7,0))/C14,1))',
                        num_fmt="0.0"),
        ],
        checks=[
            Check("وضعیت معتبرها",
                  '=IF(C15="—","",IF(C14<COUNT(A6:H7)*0.8,"❌ حذف >20% — تکرار","✅"))', 16),
            Check("Rm اصلاح‌شده (دما + زاویه)",
                  '=IF(OR(C15="—",C9="",C10="",C11=""),"—",ROUND(C15*IF(C9="مرطوب",0.95,1)*IF(C11<10,1.03,1)*IF(C10="عمودی به پایین-90",1.08,IF(C10="عمودی به بالا+90",0.92,1)),1))', 17),
        ]
    ))

define_tests()

# ─── Workbook Builder ──────────────────────────────────────────────────────
class WorkbookBuilder:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # remove default sheet
        self._named_ranges = {}

    def build(self) -> Workbook:
        """Build all sheets in order."""
        self._build_guide()
        self._build_info()
        self._build_all_tests()
        self._build_report()
        self._build_dashboard()
        self._build_qa()
        self._build_reference_sheets()
        self._protect_all()
        self._apply_global_settings()
        return self.wb

    # ─── Helper methods ──────────────────────────────────────────────────
    @staticmethod
    def _cell(ws, row, col, value=None, style=None, num_fmt=None, locked=True, hyperlink=None):
        cell = ws.cell(row=row, column=col, value=value)
        if style:
            STYLES.apply(cell, style)
        if num_fmt:
            cell.number_format = num_fmt
        cell.protection = Protection(locked=locked)
        if hyperlink:
            cell.hyperlink = hyperlink
        return cell

    @staticmethod
    def _merge(ws, r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def _add_nav(self, ws):
        for col in range(1, 12):
            cell = ws.cell(row=1, column=col)
            STYLES.apply(cell, "nav")
        self._cell(ws, 1, 1, "🏠", hyperlink="#00_راهنما!A1")
        self._cell(ws, 1, 3, "📋", hyperlink="#01_اطلاعات_آزمون!A1")
        self._cell(ws, 1, 5, "📑", hyperlink="#22_گزارش!A1")
        self._cell(ws, 1, 7, "📊", hyperlink="#23_داشبورد!A1")
        self._cell(ws, 1, 9, "🔒")

    def _add_title(self, ws, title, subtitle=""):
        self._merge(ws, 2, 1, 2, 11)
        cell = ws.cell(row=2, column=1, value=title)
        STYLES.apply(cell, "title")
        cell.alignment = STYLES.get("right")["alignment"]  # type: ignore
        if subtitle:
            self._merge(ws, 3, 1, 3, 11)
            cell2 = ws.cell(row=3, column=1, value=subtitle)
            STYLES.apply(cell2, "subtitle")
            cell2.alignment = STYLES.get("right")["alignment"]  # type: ignore

    @staticmethod
    def _add_dv(ws, cells, dv_type, **kwargs):
        if dv_type == "decimal":
            dv = DataValidation(type="decimal", operator="between",
                                formula1=str(kwargs.get("min", 0)),
                                formula2=str(kwargs.get("max", 100000)),
                                allow_blank=kwargs.get("allow_blank", True))
        elif dv_type == "list":
            dv = DataValidation(type="list", formula1=kwargs.get("formula", ""),
                                allow_blank=kwargs.get("allow_blank", True))
        elif dv_type == "custom":
            dv = DataValidation(type="custom", formula1=kwargs.get("formula", ""),
                                allow_blank=kwargs.get("allow_blank", True))
        else:
            return
        dv.error = kwargs.get("error_msg", "مقدار نامعتبر")
        dv.errorTitle = "خطای ورودی"
        dv.prompt = kwargs.get("prompt", "")
        dv.promptTitle = "راهنما"
        ws.add_data_validation(dv)
        for cell_ref in cells:
            dv.add(cell_ref)

    def _apply_input_validation(self, ws, inputs):
        for inp in inputs:
            if inp.validation:
                self._add_dv(ws, [f"{get_column_letter(inp.col)}{inp.row}"],
                             **inp.validation)
            if inp.tooltip:
                ws.cell(row=inp.row, column=inp.col).comment = Comment(inp.tooltip, "سیستم")

    @staticmethod
    def _freeze_panes(ws, row=4, col=1):
        ws.freeze_panes = f"{get_column_letter(col)}{row}"

    @staticmethod
    def _setup_print(ws, orientation="portrait", fit_to_width=1, fit_to_height=1,
                     margins=(0.3, 0.3, 0.5, 0.5)):
        ws.page_setup.orientation = orientation
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = fit_to_width
        ws.page_setup.fitToHeight = fit_to_height
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(
            left=margins[0], right=margins[0],
            top=margins[2], bottom=margins[3]
        )

    # ─── Build specific sheets ───────────────────────────────────────────
    def _build_guide(self):
        ws = self.wb.create_sheet("00_راهنما")
        ws.sheet_properties.tabColor = "1F4E79"
        ws.sheet_view.rightToLeft = True
        self._add_nav(ws)
        self._add_title(ws, "🧪 همراه دیجیتال آزمایشگاه فناوری بتن",
                        f"نسخه {VERSION} | ساخت: {BUILD_DATE}")
        r = 5
        self._cell(ws, r, 1, "📌 راهنمای سریع", style="title", align="right")
        self._merge(ws, r, 1, r, 11)
        r += 1
        self._cell(ws, r, 1, "سلول‌های زرد = ورودی کاربر | خاکستری = محاسبه خودکار | سبز = قبول | نارنجی = هشدار | قرمز = خطا",
                   style="normal", align="right")
        self._merge(ws, r, 1, r, 11)
        r += 2
        self._cell(ws, r, 1, "📋 فهرست شیت‌ها:", style="title", align="right")
        self._merge(ws, r, 1, r, 11)
        r += 1
        for code, name, std in [
            ("01", "اطلاعات آزمون", ""),
            ("02", "۱-۱ دانه‌بندی", STANDARDS["C136"]["name"]),
            ("03", "۱-۲ رطوبت", STANDARDS["C566"]["name"]),
            ("04", "۱-۳ چگالی درشت", STANDARDS["C127"]["name"]),
            ("05", "۱-۴ چگالی ریز", STANDARDS["C128"]["name"]),
            ("06", "۱-۵ وزن واحد سنگدانه", STANDARDS["C29"]["name"]),
            ("07", "۱-۶ معادل ماسه", STANDARDS["D2419"]["name"]),
            ("08", "۱-۷ شاخص شکل", STANDARDS["D4791"]["name"]),
            ("09", "۱-۸ جذب آب", STANDARDS["C127"]["name"]),
            ("10", "۲-۱ چگالی بتن تازه", STANDARDS["C138"]["name"]),
            ("11", "۲-۲ قوام نرمال", STANDARDS["C187"]["name"]),
            ("12", "۲-۳ زمان گیرش", STANDARDS["C191"]["name"]),
            ("13", "۲-۴ مقاومت ملات", STANDARDS["EN196-1"]["name"]),
            ("14", "۳-۱ اسلامپ", STANDARDS["C143"]["name"]),
            ("15", "۳-۲ آب‌انداختگی", STANDARDS["C232"]["name"]),
            ("16", "۳-۳ وزن واحد بتن", STANDARDS["C138"]["name"]),
            ("17", "۴-۱ مقاومت فشاری", STANDARDS["C39"]["name"]),
            ("18", "۴-۲ مقاومت کششی", STANDARDS["C496"]["name"]),
            ("19", "۴-۳ مقاومت خمشی", STANDARDS["C78"]["name"] + "/" + STANDARDS["C293"]["name"]),
            ("20", "۴-۴ اولتراسونیک", STANDARDS["C597"]["name"]),
            ("21", "۴-۵ چکش اشمیت", STANDARDS["C805"]["name"]),
            ("22", "گزارش", ""),
            ("23", "داشبورد", ""),
            ("24", "QA Test", ""),
        ]:
            self._cell(ws, r, 1, code, style="num")
            self._cell(ws, r, 2, name, style="normal", align="right")
            self._cell(ws, r, 6, std, style="num")
            self._merge(ws, r, 2, r, 5)
            self._merge(ws, r, 6, r, 11)
            r += 1
        self._setup_print(ws, orientation="landscape")
        self._freeze_panes(ws, row=4)

    def _build_info(self):
        ws = self.wb.create_sheet("01_اطلاعات_آزمون")
        ws.sheet_properties.tabColor = "2E7D32"
        ws.sheet_view.rightToLeft = True
        self._add_nav(ws)
        self._add_title(ws, "📋 اطلاعات آزمون", "این اطلاعات در گزارش تکرار می‌شود")
        fields = ["نام پروژه", "شماره نمونه", "تاریخ آزمون", "نام اپراتور",
                  "دستگاه / تجهیزات", "دمای محیط (°C)", "رطوبت نسبی (%)", "استاندارد مرجع", "توضیحات"]
        r = 5
        for label in fields:
            self._cell(ws, r, 1, label, style="normal", align="right")
            self._merge(ws, r, 1, r, 2)
            self._cell(ws, r, 3, None, style="input", locked=False, align="right")
            self._merge(ws, r, 3, r, 6)
            r += 1
        self._add_dv(ws, [f"C{5+i}" for i in range(len(fields))],
                     "textLength", min_val=0, allow_blank=False,
                     error_msg="تکمیل این فیلد الزامی است", operator="greaterThan")
        self._setup_print(ws)
        self._freeze_panes(ws, row=4)

    def _build_all_tests(self):
        for test_id, spec in TEST_REGISTRY.items():
            self._build_test_sheet(spec)

    def _build_test_sheet(self, spec: TestSpec):
        ws = self.wb.create_sheet(spec.sheet_name)
        ws.sheet_properties.tabColor = spec.tab_color
        ws.sheet_view.rightToLeft = True
        self._add_nav(ws)
        self._add_title(ws, spec.title, f"{spec.standard}")

        # Inputs
        for inp in spec.inputs:
            self._cell(ws, inp.row, 1, inp.label, style="normal", align="right")
            self._merge(ws, inp.row, 1, inp.row, 2)
            self._cell(ws, inp.row, inp.col, None, style="input", locked=False, align="right")
            if inp.unit:
                self._cell(ws, inp.row, inp.col + 1, f"[{inp.unit}]", style="num")
        self._apply_input_validation(ws, spec.inputs)

        # Outputs
        for out in spec.outputs:
            self._cell(ws, out.row, 1, out.label, style="normal", align="right")
            self._merge(ws, out.row, 1, out.row, 2)
            self._cell(ws, out.row, out.col, out.formula, style=out.style,
                       num_fmt=out.num_fmt, locked=True)

        # Checks
        for chk in spec.checks:
            self._cell(ws, chk.row, 1, chk.label, style="normal", align="right")
            self._merge(ws, chk.row, 1, chk.row, 2)
            self._cell(ws, chk.row, chk.col, chk.formula, style="calc", locked=True)

        # Chart if defined
        if spec.chart:
            self._add_chart(ws, spec)

        # Conditional formatting for status cells
        self._add_status_cf(ws, spec)

        self._setup_print(ws)
        self._freeze_panes(ws, row=4)

    def _add_chart(self, ws, spec):
        chart = ScatterChart()
        chart.title = spec.chart.get("title", "")
        chart.style = 13
        chart.x_axis.title = "اندازه الک (mm)"
        chart.y_axis.title = "% عبوری"
        chart.x_axis.scaling.logBase = 10
        chart.x_axis.scaling.min = 0.075
        chart.x_axis.scaling.max = 75
        chart.width = 20
        chart.height = 12

        data_start = spec.chart["data_start"]
        data_end = spec.chart["data_end"]
        x_col = spec.chart["x_col"]
        y_col = spec.chart["y_col"]

        xvalues = Reference(ws, min_col=x_col, min_row=data_start, max_row=data_end)
        yvalues = Reference(ws, min_col=y_col, min_row=data_start, max_row=data_end)
        series = Series(yvalues, xvalues, title="نمونه")
        series.graphicalProperties.line.solidFill = "1F4E79"
        series.graphicalProperties.line.width = 22500
        chart.series.append(series)

        # Upper/lower limits (hardcoded for sieve chart)
        if spec.id == "1-1":
            upper = Reference(ws, min_col=3, min_row=data_start, max_row=data_end)
            lower = Reference(ws, min_col=4, min_row=data_start, max_row=data_end)
            for ref, title, color in [(upper, "حد بالا", "ED7D31"), (lower, "حد پایین", "ED7D31")]:
                s = Series(ref, xvalues, title=title)
                s.graphicalProperties.line.solidFill = color
                s.graphicalProperties.line.dashStyle = "dash"
                chart.series.append(s)

        ws.add_chart(chart, f"H{data_start-1}")

    def _add_status_cf(self, ws, spec):
        # Apply conditional formatting to check cells (col 3)
        for chk in spec.checks:
            cell_ref = f"{get_column_letter(chk.col)}{chk.row}"
            # If cell contains "✅", fill green
            ws.conditional_formatting.add(
                cell_ref,
                FormulaRule(formula=[f'ISNUMBER(FIND("✅", {cell_ref}))'],
                            fill=PatternFill("solid", fgColor=STYLES.COLORS["pass_fill"]),
                            font=Font(color=STYLES.COLORS["pass_font"]))
            )
            # If contains "❌" or "⚠️", fill red/orange
            ws.conditional_formatting.add(
                cell_ref,
                FormulaRule(formula=[f'ISNUMBER(FIND("❌", {cell_ref}))'],
                            fill=PatternFill("solid", fgColor=STYLES.COLORS["fail_fill"]),
                            font=Font(color=STYLES.COLORS["fail_font"]))
            )
            ws.conditional_formatting.add(
                cell_ref,
                FormulaRule(formula=[f'ISNUMBER(FIND("⚠️", {cell_ref}))'],
                            fill=PatternFill("solid", fgColor=STYLES.COLORS["warn_fill"]),
                            font=Font(color=STYLES.COLORS["warn_font"]))
            )

    # ─── Report ──────────────────────────────────────────────────────────
    def _build_report(self):
        ws = self.wb.create_sheet("22_گزارش")
        ws.sheet_properties.tabColor = "9C27B0"
        ws.sheet_view.rightToLeft = True
        self._add_nav(ws)
        self._add_title(ws, "📑 گزارش آزمایشگاهی", "خلاصه نتایج")
        r = 5

        # Project info
        self._cell(ws, r, 1, "اطلاعات پروژه:", style="title", align="right")
        self._merge(ws, r, 1, r, 11)
        r += 1
        for i, label in enumerate(["نام پروژه", "شماره نمونه", "تاریخ", "اپراتور", "استاندارد"]):
            self._cell(ws, r, 1, label, style="normal", align="right")
            self._cell(ws, r, 2, f"='01_اطلاعات_آزمون'!C{5+i}", style="calc", align="right")
            self._merge(ws, r, 2, r, 6)
            r += 1

        r += 1
        self._cell(ws, r, 1, "خلاصه نتایج:", style="title", align="right")
        self._merge(ws, r, 1, r, 11)
        r += 1
        headers = ["آزمایش", "نتیجه", "واحد", "وضعیت"]
        for col, h in enumerate(headers, 1):
            self._cell(ws, r, col, h, style="header", align="center")

        # Result mapping (using registry)
        result_map = {
            "۱-۱ دانه‌بندی (FM)": ("02_آزمایش_1-1", "C19", ""),
            "۱-۲ رطوبت": ("03_آزمایش_1-2", "C8", "%"),
            "۱-۳ چگالی درشت (SSD)": ("04_آزمایش_1-3", "C10", ""),
            "۱-۴ چگالی ریز (SSD)": ("05_آزمایش_1-4", "C13", ""),
            "۱-۵ وزن واحد حجمی": ("06_آزمایش_1-5", "C9", "kg/m³"),
            "۱-۶ معادل ماسه": ("07_آزمایش_1-6", "C8", "%"),
            "۱-۷ شاخص درازگی": ("08_آزمایش_1-7", "C10", "%"),
            "۱-۸ جذب آب": ("09_آزمایش_1-8", "C8", "%"),
            "۲-۱ چگالی بتن تازه": ("10_آزمایش_2-1", "C9", "kg/m³"),
            "۲-۲ قوام نرمال (w/c)": ("11_آزمایش_2-2", "C9", "%"),
            "۲-۳ گیرش اولیه": ("12_آزمایش_2-3", "C11", "min"),
            "۲-۴ مقاومت ملات": ("13_آزمایش_2-4", "D12", "MPa"),
            "۳-۱ اسلامپ": ("14_آزمایش_3-1", "C8", "mm"),
            "۳-۲ آب‌انداختگی": ("15_آزمایش_3-2", "C9", "%"),
            "۳-۳ وزن واحد بتن": ("16_آزمایش_3-3", "C10", "kg/m³"),
            "۴-۱ مقاومت فشاری": ("17_آزمایش_4-1", "C11", "MPa"),
            "۴-۲ مقاومت کششی": ("18_آزمایش_4-2", "C9", "MPa"),
            "۴-۳ مقاومت خمشی": ("19_آزمایش_4-3", "C13", "MPa"),
            "۴-۴ سرعت پالس": ("20_آزمایش_4-4", "C8", "km/s"),
            "۴-۵ چکش اشمیت (Rm)": ("21_آزمایش_4-5", "C15", ""),
        }

        r += 1
        for name, (sheet, cell, unit) in result_map.items():
            self._cell(ws, r, 1, name, style="normal", align="right")
            self._cell(ws, r, 2, f"='{sheet}'!{cell}", style="calc", num_fmt="0.0")
            self._cell(ws, r, 3, unit, style="num")
            status_formula = f'=IF(ISNUMBER(B{r}),"✅",IF(B{r}="—","انجام نشده",B{r}))'
            self._cell(ws, r, 4, status_formula, style="calc")
            # CF for status
            ws.conditional_formatting.add(
                f"D{r}",
                FormulaRule(formula=[f'ISNUMBER(FIND("✅", D{r}))'],
                            fill=PatternFill("solid", fgColor=STYLES.COLORS["pass_fill"]),
                            font=Font(color=STYLES.COLORS["pass_font"]))
            )
            r += 1

        self._setup_print(ws, orientation="landscape")
        self._freeze_panes(ws, row=4)

    # ─── Dashboard ────────────────────────────────────────────────────────
    def _build_dashboard(self):
        ws = self.wb.create_sheet("23_داشبورد")
        ws.sheet_properties.tabColor = "00BCD4"
        ws.sheet_view.rightToLeft = True
        self._add_nav(ws)
        self._add_title(ws, "📊 داشبورد پروژه", "وضعیت کلی")
        r = 5

        self._cell(ws, r, 1, "تعداد کل آزمایش‌ها:", style="normal", align="right")
        self._cell(ws, r, 3, len(TEST_REGISTRY), style="calc", num_fmt="0")
        r += 1

        self._cell(ws, r, 1, "شیت‌های پیاده‌سازی‌شده:", style="normal", align="right")
        self._cell(ws, r, 3, len(TEST_REGISTRY), style="calc", num_fmt="0")
        r += 1

        # Progress calculation
        input_ranges = []
        for spec in TEST_REGISTRY.values():
            for inp in spec.inputs:
                input_ranges.append(f"'{spec.sheet_name}'!{get_column_letter(inp.col)}{inp.row}")
        if input_ranges:
            range_str = ",".join(input_ranges)
            total_cells = len(input_ranges)
            progress_formula = f'=IFERROR(COUNTA({range_str})/{total_cells},0)'
            self._cell(ws, r, 1, "درصد پیشرفت (ورودی‌های پرشده):", style="normal", align="right")
            self._cell(ws, r, 3, progress_formula, style="calc", num_fmt="0%")
        r += 2

        self._cell(ws, r, 1, "نسخه:", style="normal", align="right")
        self._cell(ws, r, 3, VERSION, style="calc")
        r += 1
        self._cell(ws, r, 1, "تاریخ ساخت:", style="normal", align="right")
        self._cell(ws, r, 3, BUILD_DATE, style="calc")

        self._setup_print(ws)
        self._freeze_panes(ws, row=4)

    # ─── QA Test ──────────────────────────────────────────────────────────
    def _build_qa(self):
        ws = self.wb.create_sheet("24_QA_Test")
        ws.sheet_properties.tabColor = "FF9800"
        ws.sheet_view.rightToLeft = True
        self._add_nav(ws)
        self._add_title(ws, "🧪 QA Test — تست‌های خودکار", "اعتبارسنجی فرمول‌ها")
        r = 5
        headers = ["Test ID", "شرح", "ورودی", "انتظار", "نتیجه", "وضعیت"]
        for col, h in enumerate(headers, 1):
            self._cell(ws, r, col, h, style="header", align="center")
        r += 1

        tests = [
            ("T-001", "ورودی خالی", "همه سلول‌ها پاک", "بدون #DIV/0!",
             '=IF(COUNTA(\'03_آزمایش_1-2\'!C5:C6)=0,"✅ PASS","—")'),
            ("T-002", "مرزی: W2=0", "W1=100, W2=0", "نمایش '—'",
             '=IF(\'03_آزمایش_1-2\'!C8="—","✅ PASS","❌ FAIL")'),
            ("T-003", "نمونه کتاب ۲-۴", "3340 kgf, A=1600", "≈20.5 MPa",
             '=IF(ABS(\'13_آزمایش_2-4\'!D12-20.5)<0.2,"✅ PASS","❌ FAIL")'),
            ("T-004", "نمونه کتاب ۴-۱", "d=150, F=715 kN", "≈40.4 MPa",
             '=IF(ABS(\'17_آزمایش_4-1\'!C11-40.4)<0.2,"✅ PASS","❌ FAIL")'),
        ]

        for test_id, desc, inp, expected, formula in tests:
            self._cell(ws, r, 1, test_id, style="num")
            self._cell(ws, r, 2, desc, style="normal", align="right")
            self._cell(ws, r, 3, inp, style="normal", align="right")
            self._cell(ws, r, 4, expected, style="normal", align="right")
            self._cell(ws, r, 5, formula, style="calc")
            status_cell = ws.cell(row=r, column=6)
            status_cell.value = f'=IF(ISNUMBER(FIND("PASS",E{r})),"✅",IF(ISNUMBER(FIND("FAIL",E{r})),"❌","—"))'
            status_cell.font = Font(name=STYLES.FONT_BODY, size=11)
            status_cell.alignment = STYLES.get("center")["alignment"]  # type: ignore
            r += 1

        self._setup_print(ws)
        self._freeze_panes(ws, row=4)

    # ─── Reference Sheets ────────────────────────────────────────────────
    def _build_reference_sheets(self):
        # _Reference_DB
        ws = self.wb.create_sheet("_Reference_DB")
        ws.sheet_state = "hidden"
        ws.append(["Test ID", "Parameter", "Unit", "Standard", "Tolerance"])
        for spec in TEST_REGISTRY.values():
            for out in spec.outputs:
                ws.append([spec.id, out.label, out.unit, spec.standard, ""])

        # _Standards
        ws = self.wb.create_sheet("_Standards")
        ws.sheet_state = "hidden"
        ws.append(["ID", "Standard", "Edition", "Title"])
        for key, std in STANDARDS.items():
            ws.append([key, std["name"], std["edition"], std["title"]])

        # _Validation_Data
        ws = self.wb.create_sheet("_Validation_Data")
        ws.sheet_state = "hidden"
        ws.append(["Test ID", "Case", "Input", "Expected", "Tolerance"])

        # _Glossary
        ws = self.wb.create_sheet("_Glossary")
        ws.sheet_state = "hidden"
        ws.append(["Persian", "English", "Definition"])

        # _Materials_DB
        ws = self.wb.create_sheet("_Materials_DB")
        ws.sheet_state = "hidden"
        ws.append(["Material", "Property", "Value", "Unit"])

    # ─── Protection & Global Settings ──────────────────────────────────
    def _protect_all(self):
        for ws in self.wb.worksheets:
            if ws.title.startswith("_"):
                continue
            ws.protection.sheet = True
            ws.protection.password = PASSWORD
            ws.protection.selectLockedCells = False
            ws.protection.selectUnlockedCells = True
            ws.protection.formatCells = False
            ws.protection.formatColumns = False
            ws.protection.formatRows = False
            ws.protection.insertColumns = False
            ws.protection.insertRows = False
            ws.protection.deleteColumns = False
            ws.protection.deleteRows = False
        self.wb.security.lockStructure = True
        self.wb.security.workbookPassword = PASSWORD

    def _apply_global_settings(self):
        for ws in self.wb.worksheets:
            # Set column widths
            for col in range(1, 12):
                ws.column_dimensions[get_column_letter(col)].width = 14
            # Zoom
            ws.sheet_view.zoomScale = 80

# ─── Main ──────────────────────────────────────────────────────────────────
def compute_sha256(filepath: Path) -> str:
    sha256 = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)
    return sha256.hexdigest()

def main():
    parser = argparse.ArgumentParser(description="Concrete Lab Companion Generator")
    parser.add_argument("--output", default="output", help="Output directory")
    parser.add_argument("--no-protect", action="store_true", help="Disable sheet protection")
    parser.add_argument("--version", action="version", version=f"v{VERSION}")
    args = parser.parse_args()

    logger.info("═" * 50)
    logger.info(f" Concrete Lab Companion Generator v{VERSION}")
    logger.info("═" * 50)

    builder = WorkbookBuilder()
    wb = builder.build()

    output_dir = Path(args.output)
    output_dir.mkdir(exist_ok=True)
    filename = f"Concrete_Lab_Companion_v{VERSION}.xlsx"
    output_path = output_dir / filename

    if output_path.exists():
        try:
            output_path.unlink()
        except PermissionError:
            logger.error("فایل قبلی باز است. لطفاً آن را ببندید و دوباره اجرا کنید.")
            sys.exit(1)

    wb.save(str(output_path))
    logger.info(f"💾 Saved: {output_path}")

    sha = compute_sha256(output_path)
    sha_path = output_path.parent / (output_path.name + ".sha256")
    with open(sha_path, "w", encoding="utf-8") as f:
        f.write(f"{sha}  {filename}\n")
    logger.info(f"🔐 SHA-256: {sha}")

    logger.info("✅ BUILD COMPLETE")
    logger.info("═" * 50)

if __name__ == "__main__":
    main()
