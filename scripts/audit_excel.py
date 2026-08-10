import sys, os, json, re, glob
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import load_workbook

CYAN = "\033[96m"; GREEN = "\033[92m"; YELLOW = "\033[93m"; RED = "\033[91m"; RESET = "\033[0m"
def log(msg, color=""): print(f"{color}{msg}{RESET}")

def main():
    # پیدا کردن فایل در .audit/
    candidates = glob.glob(".audit/*.xlsx")
    if not candidates:
        log("❌ فایل اکسل در .audit/ یافت نشد!", RED)
        return 1
    wb_path = max(candidates, key=os.path.getmtime)

    log("\n" + "="*70, CYAN)
    log("🔬 EXCEL OBSESSIVE AUDIT - روی فایل Release", CYAN)
    log(f"📁 فایل: {wb_path}", GREEN)
    log(f"📏 حجم: {os.path.getsize(wb_path)/1024:.1f} KB", GREEN)
    log("="*70, CYAN)

    wb = load_workbook(wb_path, data_only=False)

    issues = []
    positives = []
    formula_lines = [
        f"# Excel Formula Audit - {os.path.basename(wb_path)}",
        f"# Downloaded from GitHub Release",
        f"# Sheets: {len(wb.sheetnames)}",
        ""
    ]

    volatile_funcs = ["INDIRECT", "OFFSET", "NOW", "TODAY", "RAND", "RANDBETWEEN", "INFO", "AREAS"]
    circular_refs = []
    hardcoded_refs = 0
    named_ranges_count = len(list(wb.defined_names.definedName)) if hasattr(wb, 'defined_names') and wb.defined_names else 0

    # منتقد ۲: بررسی نام شیت‌ها (Excel Online)
    dash_sheets = [s for s in wb.sheetnames if "-" in s and not s.startswith("_")]
    if dash_sheets:
        issues.append(f"نام شیت‌ها دارای '-' است (مشکل Excel Online): {dash_sheets[:3]}...")
    else:
        positives.append("نام شیت‌ها استاندارد است ✅")

    # منتقد ۵: Named Ranges
    if named_ranges_count < 15:
        issues.append(f"تعداد Named Ranges کم است ({named_ranges_count}). خطر ارجاعات سخت‌افزاری.")
    else:
        positives.append(f"Named Ranges: {named_ranges_count} ✅")

    # منتقد ۲: Calculation Mode
    calc_mode = "unknown"
    if hasattr(wb, 'calculation') and wb.calculation:
        if hasattr(wb.calculation, 'calcMode'):
            calc_mode = wb.calculation.calcMode
    if calc_mode != "auto" and calc_mode != "autoNoTable":
        issues.append(f"Calculation Mode: '{calc_mode}' (باید auto باشد)")
    else:
        positives.append(f"Calculation Mode: {calc_mode} ✅")

    # استخراج فرمول‌ها
    total_formulas = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        formula_lines.append(f"## Sheet: {sheet_name}")
        sheet_formulas = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                    coord = cell.coordinate
                    formula_lines.append(f"  {coord}: {formula}")
                    total_formulas += 1
                    sheet_formulas += 1

                    clean_coord = coord.replace("$", "")
                    if clean_coord in formula.replace("$", "") and "!" not in formula.split(clean_coord)[0][-5:]:
                        circular_refs.append(f"{sheet_name}!{coord}: {formula}")
                        issues.append(f"🚨 فرمول دایره‌ای (منتقد ۸): {sheet_name}!{coord}")

                    for vf in volatile_funcs:
                        if re.search(rf"\b{vf}\s*\(", formula, re.IGNORECASE):
                            issues.append(f"⚠️ تابع فرار {vf} در {sheet_name}!{coord}")

                    if re.search(r"\b[A-Z]{1,3}\$?\d+\b", formula.replace("$","")):
                        hardcoded_refs += 1

        formula_lines.append(f"  # {sheet_formulas} formulas")
        formula_lines.append("")

    formula_lines.insert(4, f"# Total formulas: {total_formulas}")

    if hardcoded_refs > 50 and named_ranges_count < 20:
        issues.append(f"🚨 بحران ارجاعات سخت‌افزاری (منتقد ۵): {hardcoded_refs} ارجاع مستقیم")

    # ذخیره formulas_audit.txt (حل مشکل Black Box منتقد ۳)
    os.makedirs("validation", exist_ok=True)
    with open("validation/formulas_audit.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(formula_lines))
    positives.append(f"✅ {total_formulas} فرمول برای Git Diff استخراج شد")

    # Golden Tests (منتقد ۶)
    golden_dir = "validation/golden_cases"
    golden_files = glob.glob(os.path.join(golden_dir, "*.json")) if os.path.exists(golden_dir) else []
    if len(golden_files) == 0:
        issues.append("🚨 Golden Tests = 0 (منتقد ۶): هیچ تست ماشینی وجود ندارد.")
    else:
        positives.append(f"Golden Tests: {len(golden_files)} فایل تست ✅")

    # Data Validation & Conditional Formatting (منتقد ۷ و ۸)
    total_dv = 0
    total_cf = 0
    protected_sheets = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.data_validations and ws.data_validations.dataValidation:
            total_dv += len(ws.data_validations.dataValidation)
        try:
            total_cf += len(ws.conditional_formatting._cf_rules) if hasattr(ws.conditional_formatting, '_cf_rules') else len(list(ws.conditional_formatting))
        except: pass
        if ws.protection and ws.protection.sheet:
            protected_sheets += 1

    if total_dv < 10:
        issues.append(f"Data Validation ناکافی ({total_dv} قانون) - منتقد ۷")
    else:
        positives.append(f"Data Validation: {total_dv} قانون ✅")

    if total_cf == 0:
        issues.append("Conditional Formatting = 0 - منتقد ۸")
    else:
        positives.append(f"Conditional Formatting: {total_cf} قانون ✅")

    positives.append(f"شیت‌های محافظت‌شده: {protected_sheets}/{len(wb.sheetnames)}")

    # چاپ نتایج
    log("\n📊 نقاط قوت:", GREEN)
    for p in positives: log(f"  ✅ {p}", GREEN)

    log("\n🚨 ایرادات:", YELLOW)
    unique_issues = list(dict.fromkeys(issues))
    for i in unique_issues: log(f"  ❌ {i}", YELLOW)

    # گزارش JSON
    report = {
        "phase": 2,
        "source": "GitHub Release",
        "file": os.path.basename(wb_path),
        "size_kb": os.path.getsize(wb_path) / 1024,
        "sheets": len(wb.sheetnames),
        "named_ranges": named_ranges_count,
        "total_formulas": total_formulas,
        "hardcoded_refs": hardcoded_refs,
        "circular_refs": circular_refs,
        "data_validations": total_dv,
        "conditional_formatting": total_cf,
        "protected_sheets": protected_sheets,
        "golden_tests": len(golden_files),
        "calc_mode": calc_mode,
        "issues": unique_issues,
        "positives": positives
    }

    os.makedirs("docs/qa", exist_ok=True)
    with open("docs/qa/EXCEL_AUDIT_REPORT.json", "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    log(f"\n📄 گزارش JSON: docs/qa/EXCEL_AUDIT_REPORT.json", CYAN)
    log(f"📄 لیست فرمول‌ها: validation/formulas_audit.txt", CYAN)
    return 0

if __name__ == "__main__":
    sys.exit(main())
