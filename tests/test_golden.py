import openpyxl
import pytest
import os

EXCEL_FILE = "Concrete_Lab_Companion_v2.1.0.xlsx" # نام فایل نهایی

def test_mortar_strength_2_4():
    if not os.path.exists(EXCEL_FILE): pytest.skip("Excel file not found")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['2_4_Mortar'] # یا نام دقیق شیت
    # فرض: داده‌های طلایی در شیت وارد شده‌اند
    # مقاومت خمشی باید حدود 7.5 مگاپاسکال باشد
    assert ws['D13'].value == pytest.approx(7.5, rel=1e-1)

def test_density_physical_logic_1_4():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['1_4j_Weight']
    od = ws['B5'].value # فرض
    ssd = ws['B6'].value # فرض
    assert ssd >= od, "SSD must be greater than or equal to OD (Physical Law)"
