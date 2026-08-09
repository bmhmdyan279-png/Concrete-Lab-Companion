import openpyxl
import os

def dump_formulas():
    wb_path = next((f for f in os.listdir('releases') if f.endswith('.xlsx')), None)
    if not wb_path: return print("❌ No xlsx found.")
    
    wb = openpyxl.load_workbook(os.path.join('releases', wb_path), data_only=False)
    with open('validation/formulas_dump.txt', 'w', encoding='utf-8') as out:
        out.write(f"# FORMULA DUMP - {wb_path}\n")
        out.write("# This file is auto-generated for Git Diffing and PR Audits.\n\n")
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            out.write(f"=== SHEET: {sheet} ===\n")
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).startswith('='):
                        out.write(f"[{cell.coordinate}] {cell.value}\n")
            out.write("\n")
    print("✅ Formulas exported to validation/formulas_dump.txt for Git tracking.")

if __name__ == "__main__":
    dump_formulas()
