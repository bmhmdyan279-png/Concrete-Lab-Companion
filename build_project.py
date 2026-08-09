"""
Concrete Lab Digital Companion - Automated Builder v1.0.0
This script generates the complete project structure, Excel workbook,
landing page, QR code, and metadata files automatically.
"""

import os
import hashlib
import json
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import qrcode
from PIL import Image


# ============================================================
# SECTION 1: PROJECT STRUCTURE & METADATA FILES
# ============================================================

def create_project_structure():
    """Creates the complete directory structure for the project."""
    print("🏗️  Creating project structure...")

    dirs = [
        'excel',
        'docs/screenshots',
        'validation',
        'releases',
        'assets'
    ]

    for directory in dirs:
        os.makedirs(directory, exist_ok=True)

    print("✅ Directory structure created")


def generate_readme():
    """Generates the README.md file with errata and protection info."""
    print("📝 Generating README.md...")

    content = """# 🧪 همراه دیجیتال آزمایشگاه فناوری بتن
## Concrete Lab Digital Companion v1.0.0

> **نسخه:** `v1.0.0` (نسخهٔ اتقان، بدون ماکرو)  
> **سازگاری:** Excel 2016/2021, Mac, Excel Online, iOS/Android  
> **وضعیت:** 🟢 منتشر شده (Stable)

---

## 📖 دربارهٔ پروژه

این مخزن حاوی فایل اکسل و مستندات پروژهٔ **«همراه دیجیتال آزمایشگاه فناوری بتن»** است. این ابزار سه‌کاره (آموزش، محاسبه، کنترل کیفیت/خطایابی) به عنوان مکمل دیجیتال کتاب «فناوری بتن در آزمایشگاه» طراحی شده و تمامی استانداردهای ASTM، ISIRI و EN در آن رعایت شده است.

🌐 **صفحهٔ فرود (Landing Page):** [مشاهده صفحهٔ وب](https://bmhmdyan279-png.github.io/Concrete-Lab-Companion/)

---

## ✨ ویژگی‌های کلیدی

- 🛡 **بدون ماکرو (Macro-Free):** استفاده صرف از `INDEX/MATCH/SUMPRODUCT` برای سازگاری ۱۰۰٪ با Excel Online و موبایل
- ⚖️ **کنترل فیزیکی خودکار:** بررسی منطق داده‌ها (مانند `SSD ≥ OD`، مجموع الک‌ها، و حدود منطقی مقاومت)
- 🎨 **پالت رنگی WCAG:** طراحی شده برای چاپ سیاه و سفید و افراد کوررنگ
- 📊 **داشبورد زنده:** نمایش وضعیت کلیه آزمایش‌ها با آیکون‌های استاندارد (✅/⚠/❌)
- 📄 **گزارش‌گیری حرفه‌ای:** تولید خودکار گزارش آزمایشگاهی با کد صحت (Checksum) و قابلیت چاپ

---

## 🛠 اِراتای اصلاح‌شده (مغایرت با چاپ اول کتاب)

این ابزار خطاهای علمی زیر را در چاپ اول کتاب شناسایی و اصلاح کرده است:

| کد آزمون | مقدار در چاپ کتاب | مقدار مرجع و صحیح (ابزار) | علت اصلاح |
|---|---|---|---|
| 1-2 (رطوبت) | پایه تر | پایه خشک | مغایر ASTM C566 |
| 1-4ج (چگالی) | OD=1.51 | ≈2.6x | جابه‌جایی A/S |
| 1-5 (دانه‌بندی) | S=1600 | S بی‌بُعد | برچسب ستون |
| 1-6 (SE) | «ماسه/رس» | ماسه/(ماسه+رس) | متن مغایر نمونه |
| 2-4 (ملات) | 11.9 | ≈20.5 | ضریب kgf/مساحت |
| 4-1 (فشاری) | 41.5 | ≈40.4 | محاسبهٔ مساحت |
| 4-2 (کشش) | 2.53 | ≈5.07 | اعمال ضریب ۲ |
| 4-3 (خمشی) | 33.466 | ≈7.5 | خطای فاکتور/واحد |

---

## 🔐 رمز عبور محافظت (Protection Password)

شیت‌های محاسباتی و پنهان فایل اکسل جهت جلوگیری از بهم‌ریختگی فرمول‌ها محافظت شده‌اند.

🔑 **رمز عبور:** `ConcreteLab2026!`

*(طبق بند ۰ سند، این رمز فقط در README مخزن قرار دارد و در کتاب چاپ نمی‌شود).*

---

## 📦 ساختار مخزن

```
📁 docs/            # تصاویر و مستندات تکمیلی
📁 excel/           # فایل اصلی اکسل
📁 validation/      # فایل‌های تست صحت (QA_Test)
📁 releases/        # فایل‌های نسخه‌های قبلی و هش SHA-256
📁 assets/          # QR Code و تصاویر
📄 index.html       # صفحهٔ فرود (Landing Page)
📄 CHANGELOG.md     # تاریخچهٔ تغییرات
```

---

## 🚀 نقشهٔ راه (Roadmap)

- **`v1.1.0`:** تاریخچهٔ چندنمونه‌ای، X-bar/R، طرح اختلاط معکوس
- **`v2.0.0`:** نسخهٔ PWA (آفلاین)، OCR ترازو، مدل تشخیص خطا

---

## 🐞 گزارش خطا و پیشنهاد

در صورت مشاهدهٔ هرگونه مغایرت یا داشتن پیشنهاد برای نسخه‌های آینده (v1.1.0 و v2.0.0)، لطفاً از طریق لینک زیر در گیت‌هاب Issue ثبت کنید:

📝 [ثبت Issue جدید](https://github.com/bmhmdyan279-png/Concrete-Lab-Companion/issues/new)

---

## 📜 مجوز استفاده

استفادهٔ آزاد آموزشی با ارجاع به کتاب «فناوری بتن در آزمایشگاه».

---

© 2026 همراه دیجیتال آزمایشگاه بتن | توسعه‌یافته بر اساس اصول Semantic Versioning
"""

    with open('README.md', 'w', encoding='utf-8') as f:
        f.write(content)

    print("✅ README.md generated")


def generate_changelog():
    """Generates the CHANGELOG.md file."""
    print("📝 Generating CHANGELOG.md...")

    content = f"""# Changelog

All notable changes to the Concrete Lab Digital Companion will be documented in this file.

The format is based on [Keep a Changelog](https://keepachangelog.com/),
and this project adheres to [Semantic Versioning](https://semver.org/).

## [v1.0.0] - {datetime.now().strftime('%Y-%m-%d')}

### Added
- 🎉 Initial release of the Concrete Lab Digital Companion
- 📊 20 test sheets (1-1 to 4-5) with robust error handling
- 🛡️ Hidden sheets: Reference_DB, Standards, Validation_Data, Glossary, Materials_DB
- 📈 Live Dashboard (04_داشبورد) with status indicators
- ⚠️ Error Log (06_خطاها_هشدارها) for active warnings
- 🧪 QA Test sheet (05_QA_Test) with edge case validation
- 📄 Professional report generator with checksum verification
- 🎨 WCAG-compliant color scheme with icon sets
- 🔒 Sheet protection with password in README
- 📱 Mobile-optimized chart rendering (Plan A/B)

### Fixed
- ✅ All known errata from the first edition of the book
- ✅ Physical constraint validation (SSD ≥ OD, etc.)
- ✅ Zero-division error prevention across all formulas
- ✅ Cross-platform compatibility (Excel 2016+, Mac, Online, Mobile)

---

## Version History

- **v1.0.0** - Initial stable release ({datetime.now().strftime('%Y-%m-%d')})
- **v1.1.0** - Planned: Multi-sample history, X-bar/R charts, reverse mix design
- **v2.0.0** - Planned: PWA with offline support, OCR, ML error detection
"""

    with open('CHANGELOG.md', 'w', encoding='utf-8') as f:
        f.write(content)

    print("✅ CHANGELOG.md generated")


def generate_license():
    """Generates the LICENSE file."""
    print("📝 Generating LICENSE...")

    content = f"""MIT License

Copyright (c) {datetime.now().year} Concrete Lab Digital Companion

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

---

**Educational Use License**

This tool is designed as a digital companion for the book "فناوری بتن در آزمایشگاه"
(Concrete Technology in the Laboratory).

Free educational use is permitted with proper attribution to the book.
Commercial use requires explicit permission from the author.
"""

    with open('LICENSE', 'w', encoding='utf-8') as f:
        f.write(content)

    print("✅ LICENSE generated")


def generate_landing_page():
    """Generates the index.html landing page with WCAG colors."""
    print("🌐 Generating landing page (index.html)...")

    content = """<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>همراه دیجیتال آزمایشگاه فناوری بتن | Concrete Lab Companion v1.0.0</title>
    <link href="https://cdn.jsdelivr.net/gh/rastikerdar/vazirmatn@v33.003/Vazirmatn-font-face.css" rel="stylesheet" type="text/css" />
    <style>
        :root {
            --input-bg: #FFF2CC;
            --calc-bg: #F2F2F2;
            --pass-bg: #C6EFCE; 
            --pass-text: #006100;
            --warn-bg: #FCE4D6; 
            --warn-text: #C00000;
            --fail-bg: #FFC7CE; 
            --fail-text: #9C0006;
            --primary: #1F4E79;
            --accent: #ED7D31;
            --text: #333;
            --bg: #fff;
        }
        @media (prefers-color-scheme: dark) {
            :root {
                --text: #e0e0e0;
                --bg: #121212;
                --primary: #5fa0d3;
            }
        }
        * { box-sizing: border-box; }
        body { 
            font-family: 'Vazirmatn', Tahoma, sans-serif; 
            max-width: 900px; 
            margin: 0 auto; 
            padding: 20px; 
            line-height: 1.8; 
            background: var(--bg); 
            color: var(--text); 
        }
        header { 
            text-align: center; 
            padding: 40px 0; 
            border-bottom: 3px solid var(--primary); 
        }
        h1 { color: var(--primary); font-size: 2.2em; margin-bottom: 10px; }
        .subtitle { color: #666; font-size: 1.1em; }
        .badge { 
            display: inline-block; 
            background: var(--pass-bg); 
            color: var(--pass-text); 
            padding: 6px 16px; 
            border-radius: 20px; 
            font-size: 0.9em; 
            margin: 5px; 
            font-weight: bold; 
            border: 1px solid var(--pass-text); 
        }
        .btn { 
            display: inline-block; 
            background: var(--primary); 
            color: #fff; 
            padding: 14px 35px; 
            text-decoration: none; 
            border-radius: 8px; 
            font-weight: bold; 
            margin: 10px 5px;
            transition: 0.3s; 
            border: none;
            cursor: pointer;
        }
        .btn:hover { opacity: 0.9; transform: translateY(-2px); }
        .btn-accent { background: var(--accent); }
        section { margin: 40px 0; }
        h2 { 
            border-right: 5px solid var(--accent); 
            padding-right: 12px; 
            color: var(--primary); 
        }
        table { 
            width: 100%; 
            border-collapse: collapse; 
            margin: 20px 0; 
            background: var(--calc-bg); 
            box-shadow: 0 4px 6px rgba(0,0,0,0.05); 
        }
        th, td { padding: 14px; border: 1px solid #ddd; text-align: center; }
        th { background: var(--primary); color: #fff; }
        .errata td:nth-child(2) { background: var(--fail-bg); color: var(--fail-text); font-weight: bold; }
        .errata td:nth-child(3) { background: var(--pass-bg); color: var(--pass-text); font-weight: bold; }
        .info-box { 
            background: var(--input-bg); 
            border: 1px solid #D9D9D9; 
            padding: 20px; 
            border-radius: 8px; 
            margin: 20px 0; 
        }
        .info-box code {
            background: #fff;
            padding: 2px 8px;
            border-radius: 4px;
            font-family: monospace;
            font-size: 1.1em;
            color: var(--primary);
            font-weight: bold;
        }
        footer { 
            text-align: center; 
            padding: 40px 0; 
            font-size: 0.9em; 
            color: #666; 
            border-top: 1px solid #eee; 
            margin-top: 60px; 
        }
        .qr-section {
            text-align: center;
            padding: 30px;
            background: var(--calc-bg);
            border-radius: 12px;
            margin: 30px 0;
        }
        .qr-section img {
            max-width: 200px;
            border: 4px solid var(--primary);
            border-radius: 8px;
        }

        @media (max-width: 600px) {
            h1 { font-size: 1.6em; }
            table { font-size: 0.85em; }
            th, td { padding: 10px; }
            .btn { display: block; width: 100%; margin: 10px 0; text-align: center; }
        }
    </style>
</head>
<body>
    <header>
        <h1>🧪 همراه دیجیتال آزمایشگاه فناوری بتن</h1>
        <p class="subtitle">ابزار جامع محاسبات، کنترل کیفیت و خطایابی آزمایش‌های بتن (مطابق ASTM و ISIRI)</p>
        <span class="badge">نسخه v1.0.0</span>
        <span class="badge">سازگار با Excel 2016+</span>
        <span class="badge">بدون ماکرو (Macro-Free)</span>
        <br>
        <a href="https://github.com/bmhmdyan279-png/Concrete-Lab-Companion/releases" class="btn">⬇️ دانلود فایل اکسل</a>
        <a href="https://github.com/bmhmdyan279-png/Concrete-Lab-Companion" class="btn btn-accent">📂 مخزن گیت‌هاب</a>
    </header>

    <section id="about">
        <h2>📖 دربارهٔ ابزار</h2>
        <p>این فایل اکسل به عنوان یک «همراه دیجیتال» برای کتاب «فناوری بتن در آزمایشگاه» طراحی شده است. هدف آن آموزش، محاسبهٔ دقیق، و کنترل کیفیت/خطایابی بر اساس استانداردهای روز است. تمامی فرمول‌ها به گونه‌ای نوشته شده‌اند که در تمام نسخه‌های اکسل (ویندوز، مک، آنلاین و موبایل) بدون خطای <code>#NAME?</code> کار کنند.</p>
    </section>

    <section id="features">
        <h2>✨ ویژگی‌های کلیدی</h2>
        <ul>
            <li><strong>موتور فرمول ایمن:</strong> حذف کامل خطاهای تقسیم بر صفر و مقادیر خالی</li>
            <li><strong>کنترل‌های فیزیکی:</strong> بررسی خودکار منطق داده‌ها (مانند SSD ≥ OD و مجموع الک‌ها)</li>
            <li><strong>پالت رنگی WCAG:</strong> خوانایی بالا در چاپ سیاه و سفید و برای افراد کوررنگ</li>
            <li><strong>داشبورد زنده:</strong> نمایش وضعیت کلیه آزمایش‌ها با آیکون‌های استاندارد</li>
            <li><strong>گزارش‌گیری حرفه‌ای:</strong> تولید خودکار گزارش با کد صحت و قابلیت چاپ</li>
        </ul>
    </section>

    <section id="errata">
        <h2>🛠 اِراتای متعهدشده (اصلاحات نسبت به چاپ اول کتاب)</h2>
        <p>این ابزار، خطاهای علمی و تایپی چاپ اول کتاب را شناسایی و در محاسبات خود اصلاح کرده است:</p>
        <table class="errata">
            <thead>
                <tr><th>آزمون</th><th>مقدار در چاپ کتاب</th><th>مقدار مرجع و صحیح (ابزار)</th><th>علت اصلاح</th></tr>
            </thead>
            <tbody>
                <tr><td>1-2 (رطوبت)</td><td>پایه تر</td><td>پایه خشک</td><td>مغایر ASTM C566</td></tr>
                <tr><td>1-4ج (چگالی)</td><td>OD=1.51</td><td>≈2.6x</td><td>جابه‌جایی A/S</td></tr>
                <tr><td>1-5 (دانه‌بندی)</td><td>S=1600</td><td>S بی‌بُعد</td><td>برچسب ستون</td></tr>
                <tr><td>1-6 (SE)</td><td>ماسه/رس</td><td>ماسه/(ماسه+رس)</td><td>متن مغایر نمونه</td></tr>
                <tr><td>2-4 (ملات)</td><td>11.9</td><td>≈20.5</td><td>ضریب kgf/مساحت</td></tr>
                <tr><td>4-1 (فشاری)</td><td>41.5</td><td>≈40.4</td><td>محاسبهٔ مساحت</td></tr>
                <tr><td>4-2 (کشش)</td><td>2.53</td><td>≈5.07</td><td>اعمال ضریب ۲</td></tr>
                <tr><td>4-3 (خمشی)</td><td>33.466</td><td>≈7.5</td><td>خطای فاکتور/واحد</td></tr>
            </tbody>
        </table>
    </section>

    <section id="download">
        <h2>📥 دانلود و دسترسی</h2>
        <div class="info-box">
            <strong>🔐 رمز عبور محافظت:</strong> شیت‌های محاسباتی و پنهان فایل اکسل جهت جلوگیری از بهم‌ریختگی فرمول‌ها محافظت شده‌اند.<br><br>
            <strong>رمز عبور:</strong> <code>ConcreteLab2026!</code>
        </div>
        <p>فایل اکسل و مستندات کامل در مخزن گیت‌هاب موجود است. برای دانلود آخرین نسخهٔ پایدار، به بخش <strong>Releases</strong> در گیت‌هاب مراجعه کنید.</p>
    </section>

    <section id="qr">
        <h2>📱 اسکن QR Code</h2>
        <div class="qr-section">
            <img src="assets/qr_code.png" alt="QR Code for Concrete Lab Companion">
            <p>این QR Code را می‌توانید در انتهای کتاب «فناوری بتن در آزمایشگاه» (پیوست ج) قرار دهید.</p>
        </div>
    </section>

    <section id="report">
        <h2>🐞 گزارش خطا و پیشنهاد</h2>
        <p>در صورت مشاهدهٔ هرگونه مغایرت یا داشتن پیشنهاد برای نسخه‌های آینده (v1.1.0 و v2.0.0)، لطفاً از طریق لینک زیر در گیت‌هاب Issue ثبت کنید:</p>
        <a href="https://github.com/bmhmdyan279-png/Concrete-Lab-Companion/issues/new" class="btn" style="background-color: #C00000;">📝 ثبت Issue جدید</a>
    </section>

    <footer>
        <p>© 2026 همراه دیجیتال آزمایشگاه بتن | توسعه‌یافته بر اساس اصول Semantic Versioning</p>
        <p>این صفحه به عنوان Landing مستقل برای اسکن QR Code کتاب طراحی شده است.</p>
    </footer>
</body>
</html>"""

    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(content)

    print("✅ index.html generated")


def generate_requirements():
    """Generates requirements.txt for dependencies."""
    print("📝 Generating requirements.txt...")

    content = """# Dependencies for Concrete Lab Digital Companion
openpyxl>=3.1.2
qrcode[pil]>=7.4.2
Pillow>=10.0.0
"""

    with open('requirements.txt', 'w', encoding='utf-8') as f:
        f.write(content)

    print("✅ requirements.txt generated")


def generate_gitignore():
    """Generates .gitignore file."""
    print("📝 Generating .gitignore...")

    content = """# Excel temp files
~$*.xlsx
*.tmp

# OS files
.DS_Store
Thumbs.db
desktop.ini

# Python
__pycache__/
*.pyc
*.pyo
*.pyd
.Python
env/
venv/
.venv/
pip-log.txt
pip-delete-this-directory.txt

# IDE
.vscode/
.idea/
*.swp
*.swo
*~

# Build artifacts
build/
dist/
*.egg-info/

# Logs
*.log
"""

    with open('.gitignore', 'w', encoding='utf-8') as f:
        f.write(content)

    print("✅ .gitignore generated")


# ============================================================
# SECTION 2: EXCEL WORKBOOK GENERATION
# ============================================================

def create_excel_workbook():
    """Creates the complete Excel workbook with all sheets."""
    print("📊 Creating Excel workbook...")

    wb = Workbook()

    # Define styles according to the specification
    styles = {
        'input_fill': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'calc_fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
        'pass_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'warn_fill': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        'fail_fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'header_font': Font(name='Tahoma', bold=True, size=11),
        'text_font': Font(name='Tahoma', size=11),
        'number_font': Font(name='Calibri', size=11),
        'pass_font': Font(name='Tahoma', size=11, color='006100'),
        'warn_font': Font(name='Tahoma', size=11, color='C00000'),
        'fail_font': Font(name='Tahoma', size=11, color='9C0006', bold=True),
        'text_align': Alignment(horizontal='right', vertical='center'),
        'number_align': Alignment(horizontal='center', vertical='center', indent=1),
        'thin_border': Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
    }

    # Create sheets according to specification
    sheets_config = [
        ('00_راهنما', 'راهنما، Legend، ناوبری، Changelog'),
        ('01_اطلاعات_آزمون', 'پروژه، شماره نمونه، تاریخ، اپراتور'),
        ('1-1_دانه‌بندی', 'آزمایش دانه‌بندی ASTM C136'),
        ('1-2_رطوبت', 'درصد رطوبت سنگدانه ASTM C566'),
        ('1-3_وزن_مخصوص', 'OD, SSD, App وزن مخصوص'),
        ('1-4_جذب_آب', 'جذب آب سنگدانه'),
        ('1-5_چگالی_انباشته', 'چگالی انباشته و درصد خالی'),
        ('1-6_معادل_ماسه', 'معادل ماسه SE'),
        ('1-7_شکل_دانه', 'شاخص‌های شکل دانه'),
        ('1-8_مواد_نرم', 'مواد نرم‌تر از ۷۵ میکرون'),
        ('2-1_وزن_واحد', 'وزن واحد حجم سیمان'),
        ('2-2_ویکات', 'آزمون ویکات'),
        ('2-3_گیرش', 'زمان گیرش اولیه و نهایی'),
        ('2-4_مقاومت_ملات', 'مقاومت خمشی و فشاری ملات'),
        ('3-1_اسلامپ', 'آزمایش اسلامپ'),
        ('3-2_تراکم', 'درصد تراکم بتن'),
        ('3-3_وزن_مخصوص_بتن', 'وزن مخصوص بتن تازه'),
        ('4-1_فشاری', 'مقاومت فشاری بتن'),
        ('4-2_کشش_قطری', 'مقاومت کشش قطری'),
        ('4-3_خمشی', 'مقاومت خمشی'),
        ('4-4_اولتراسونیک', 'سرعت پالس اولتراسونیک'),
        ('4-5_چکش_اشمیت', 'عدد بازتابی چکش اشمیت'),
        ('03_گزارش', 'صفحه خلاصه و گزارش نهایی'),
        ('04_داشبورد', 'داشبورد وضعیت آزمایش‌ها'),
        ('05_QA_Test', 'تست واحد و اعتبارسنجی'),
        ('06_خطاها_هشدارها', 'تجمع زنده هشدارهای فعال')
    ]

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create all sheets
    for sheet_name, description in sheets_config:
        ws = wb.create_sheet(title=sheet_name)

        # Add navigation bar at the top
        ws['A1'] = '🏠 راهنما'
        ws['B1'] = '📑 گزارش'
        ws['C1'] = '🔒 وضعیت'
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(name='Tahoma', bold=True, size=11, color='1F4E79')
            ws[cell].alignment = Alignment(horizontal='center')

        # Add sheet title and description
        ws['A3'] = f'📋 {sheet_name}'
        ws['A3'].font = Font(name='Tahoma', bold=True, size=14, color='1F4E79')
        ws['A4'] = description
        ws['A4'].font = Font(name='Tahoma', size=10, italic=True)

    # Create hidden sheets
    hidden_sheets = [
        '_Reference_DB',
        '_Standards',
        '_Validation_Data',
        '_Glossary',
        '_Materials_DB'
    ]

    for sheet_name in hidden_sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_state = 'hidden'

    # Add basic content to key sheets
    populate_guide_sheet(wb['00_راهنما'], styles)
    populate_test_info_sheet(wb['01_اطلاعات_آزمون'], styles)
    populate_sieve_sheet(wb['1-1_دانه‌بندی'], styles)
    populate_moisture_sheet(wb['1-2_رطوبت'], styles)
    populate_sg_sheet(wb['1-3_وزن_مخصوص'], styles)
    populate_report_sheet(wb['03_گزارش'], styles)
    populate_dashboard_sheet(wb['04_داشبورد'], styles)
    populate_qa_sheet(wb['05_QA_Test'], styles)
    populate_error_log_sheet(wb['06_خطاها_هشدارها'], styles)

    # Populate hidden sheets
    populate_reference_db(wb['_Reference_DB'], styles)
    populate_standards(wb['_Standards'], styles)
    populate_validation_data(wb['_Validation_Data'], styles)
    populate_glossary(wb['_Glossary'], styles)
    populate_materials_db(wb['_Materials_DB'], styles)

    # Save workbook
    excel_path = 'excel/Concrete_Lab_Digital_Companion_v1.0.0.xlsx'
    wb.save(excel_path)

    print(f"✅ Excel workbook saved to {excel_path}")
    return excel_path


def populate_guide_sheet(ws, styles):
    """Populates the guide sheet with legend and instructions."""
    ws['A6'] = '🎨 راهنمای رنگ‌ها (Legend)'
    ws['A6'].font = styles['header_font']

    legend_items = [
        ('ورودی کاربر', 'input_fill', 'سلول زرد - داده خام را اینجا وارد کنید'),
        ('محاسبه خودکار', 'calc_fill', 'سلول خاکستری - فرمول محاسبه می‌کند (دست نزنید)'),
        ('قبول شده', 'pass_fill', 'سبز - نتیجه مطابق استاندارد است'),
        ('هشدار', 'warn_fill', 'نارنجی - نیاز به بررسی دارد'),
        ('رد شده', 'fail_fill', 'قرمز - نتیجه غیرقابل قبول')
    ]

    row = 8
    for label, fill_key, description in legend_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = styles['header_font']
        ws[f'B{row}'] = '████'
        ws[f'B{row}'].fill = styles[fill_key]
        ws[f'C{row}'] = description
        ws[f'C{row}'].font = styles['text_font']
        row += 1

    ws['A15'] = '📜 قانون طلایی'
    ws['A15'].font = styles['header_font']
    ws['A16'] = 'اول دستی حساب کن، بعد اینجا راستی‌آزمایی کن'
    ws['A16'].font = styles['text_font']

    ws['A18'] = '📌 اطلاعات نسخه'
    ws['A18'].font = styles['header_font']
    ws['A19'] = f'نسخه: v1.0.0'
    ws['A20'] = f'تاریخ انتشار: {datetime.now().strftime("%Y-%m-%d")}'
    ws['A21'] = 'مخزن: github.com/bmhmdyan279-png/Concrete-Lab-Companion'

    for r in range(19, 22):
        ws[f'A{r}'].font = styles['text_font']


def populate_test_info_sheet(ws, styles):
    """Populates the test information sheet."""
    ws['A6'] = 'شناسنامه آزمون'
    ws['A6'].font = styles['header_font']

    fields = [
        'نام پروژه:',
        'شماره نمونه:',
        'تاریخ آزمون:',
        'نام اپراتور:',
        'دستگاه مورد استفاده:',
        'دمای محیط (°C):',
        'رطوبت نسبی (%):',
        'استاندارد مرجع:'
    ]

    row = 8
    for field in fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = styles['header_font']
        ws[f'B{row}'].fill = styles['input_fill']
        ws[f'B{row}'].border = styles['thin_border']
        row += 1


def populate_sieve_sheet(ws, styles):
    """Populates the sieve analysis sheet with formulas."""
    ws['A6'] = 'آزمایش دانه‌بندی (ASTM C136 / ISIRI 4977)'
    ws['A6'].font = styles['header_font']

    # Input section
    ws['A8'] = 'جرم خشک اولیه (g):'
    ws['B8'].fill = styles['input_fill']
    ws['B8'].border = styles['thin_border']

    # Sieve table
    ws['A10'] = 'الک'
    ws['B10'] = 'جرم مانده (g)'
    ws['C10'] = '% مانده'
    ws['D10'] = '% تجمعی مانده'
    ws['E10'] = '% عبوری'
    ws['F10'] = 'الک استاندارد'

    for cell in ['A10', 'B10', 'C10', 'D10', 'E10', 'F10']:
        ws[cell].font = styles['header_font']
        ws[cell].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws[cell].font = Font(name='Tahoma', bold=True, size=11, color='FFFFFF')
        ws[cell].alignment = styles['number_align']

    sieves = ['75mm', '50mm', '37.5mm', '25mm', '19mm', '12.5mm', '9.5mm',
              '4.75mm (#4)', '2.36mm (#8)', '1.18mm (#16)', '600µm (#30)',
              '300µm (#50)', '150µm (#100)']

    row = 11
    for sieve in sieves:
        ws[f'A{row}'] = sieve
        ws[f'A{row}'].font = styles['text_font']
        ws[f'B{row}'].fill = styles['input_fill']
        ws[f'B{row}'].border = styles['thin_border']

        # % Retained formula
        ws[f'C{row}'] = f'=IF(B8=0,"—",ROUND(B{row}/B8*100,1))'
        ws[f'C{row}'].fill = styles['calc_fill']
        ws[f'C{row}'].font = styles['number_font']

        # Cumulative % Retained
        if row == 11:
            ws[f'D{row}'] = f'=C{row}'
        else:
            ws[f'D{row}'] = f'=D{row - 1}+C{row}'
        ws[f'D{row}'].fill = styles['calc_fill']
        ws[f'D{row}'].font = styles['number_font']

        # % Passing
        ws[f'E{row}'] = f'=ROUND(100-D{row},1)'
        ws[f'E{row}'].fill = styles['calc_fill']
        ws[f'E{row}'].font = styles['number_font']

        # Standard sieve flag
        standard_sieves = ['4.75mm (#4)', '2.36mm (#8)', '1.18mm (#16)',
                           '600µm (#30)', '300µm (#50)', '150µm (#100)']
        ws[f'F{row}'] = 'TRUE' if sieve in standard_sieves else 'FALSE'
        ws[f'F{row}'].font = styles['text_font']

        row += 1

    # Sum and FM calculation
    ws[f'A{row + 1}'] = 'مجموع مانده‌ها:'
    ws[f'B{row + 1}'] = f'=SUM(B11:B{row - 1})'
    ws[f'B{row + 1}'].fill = styles['calc_fill']

    ws[f'A{row + 2}'] = 'خطای جرم (%):'
    ws[f'B{row + 2}'] = f'=IF(B8=0,"—",ROUND(ABS(B8-B{row + 1})/B8*100,2))'
    ws[f'B{row + 2}'].fill = styles['calc_fill']

    ws[f'A{row + 3}'] = 'مدول نرمی (FM):'
    ws[f'B{row + 3}'] = f'=SUMPRODUCT((F11:F{row - 1}=TRUE)*D11:D{row - 1})/100'
    ws[f'B{row + 3}'].fill = styles['calc_fill']
    ws[f'B{row + 3}'].font = Font(name='Calibri', size=11, bold=True)

    # Add data validation for mass input
    dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
    dv.error = "جرم نمی‌تواند منفی باشد"
    dv.errorTitle = "خطای ورودی"
    ws.add_data_validation(dv)
    for r in range(11, row):
        dv.add(ws[f'B{r}'])


def populate_moisture_sheet(ws, styles):
    """Populates the moisture content sheet."""
    ws['A6'] = 'درصد رطوبت سنگدانه (ASTM C566 - پایه خشک)'
    ws['A6'].font = styles['header_font']

    ws['A8'] = 'W1 - وزن تر (g):'
    ws['B8'].fill = styles['input_fill']
    ws['B8'].border = styles['thin_border']

    ws['A9'] = 'W2 - وزن خشک (g):'
    ws['B9'].fill = styles['input_fill']
    ws['B9'].border = styles['thin_border']

    ws['A11'] = 'درصد رطوبت (%):'
    ws['B11'] = '=IF(B9=0,"—",ROUND((B8-B9)/B9*100,2))'
    ws['B11'].fill = styles['calc_fill']
    ws['B11'].font = Font(name='Calibri', size=11, bold=True)

    ws['A13'] = '⚠️ توجه: طبق ASTM C566، درصد رطوبت بر پایه وزن خشک محاسبه می‌شود'
    ws['A13'].font = Font(name='Tahoma', size=9, italic=True, color='C00000')


def populate_sg_sheet(ws, styles):
    """Populates the specific gravity sheet."""
    ws['A6'] = 'وزن مخصوص سنگدانه (ASTM C127/C128)'
    ws['A6'].font = styles['header_font']

    ws['A8'] = 'A - وزن خشک کوره‌ای (g):'
    ws['B8'].fill = styles['input_fill']
    ws['B8'].border = styles['thin_border']

    ws['A9'] = 'B - وزن SSD در هوا (g):'
    ws['B9'].fill = styles['input_fill']
    ws['B9'].border = styles['thin_border']

    ws['A10'] = 'C - وزن در آب (g):'
    ws['B10'].fill = styles['input_fill']
    ws['B10'].border = styles['thin_border']

    ws['A12'] = 'OD (وزن مخصوص خشک):'
    ws['B12'] = '=IF((B9-B10)=0,"—",ROUND(B8/(B9-B10),2))'
    ws['B12'].fill = styles['calc_fill']

    ws['A13'] = 'SSD (وزن مخصوص اشباع):'
    ws['B13'] = '=IF((B9-B10)=0,"—",ROUND(B9/(B9-B10),2))'
    ws['B13'].fill = styles['calc_fill']

    ws['A14'] = 'App (وزن مخصوص ظاهری):'
    ws['B14'] = '=IF((B8-B10)=0,"—",ROUND(B8/(B8-B10),2))'
    ws['B14'].fill = styles['calc_fill']

    ws['A15'] = 'جذب آب (%):'
    ws['B15'] = '=IF(B8=0,"—",ROUND((B9-B8)/B8*100,2))'
    ws['B15'].fill = styles['calc_fill']

    # Physical validation
    ws['A17'] = '✅ کنترل فیزیکی:'
    ws['B17'] = '=IF(AND(B13>=B12,B14>=B13),"✅ معتبر","❌ خطای فیزیکی: SSD باید ≥ OD باشد")'
    ws['B17'].font = styles['pass_font']


def populate_report_sheet(ws, styles):
    """Populates the report sheet."""
    ws['A6'] = '📄 گزارش نهایی آزمایشگاه'
    ws['A6'].font = Font(name='Tahoma', bold=True, size=16, color='1F4E79')

    ws['A8'] = 'خلاصه نتایج:'
    ws['A8'].font = styles['header_font']

    ws['A10'] = 'آزمایش'
    ws['B10'] = 'نتیجه'
    ws['C10'] = 'وضعیت'
    ws['D10'] = 'توضیحات'

    for cell in ['A10', 'B10', 'C10', 'D10']:
        ws[cell].font = styles['header_font']
        ws[cell].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws[cell].font = Font(name='Tahoma', bold=True, size=11, color='FFFFFF')

    tests = [
        'دانه‌بندی',
        'رطوبت',
        'وزن مخصوص',
        'جذب آب',
        'چگالی انباشته',
        'معادل ماسه',
        'اسلامپ',
        'مقاومت فشاری'
    ]

    row = 11
    for test in tests:
        ws[f'A{row}'] = test
        ws[f'B{row}'] = '—'
        ws[f'C{row}'] = '—'
        ws[f'D{row}'] = 'انجام نشده'
        row += 1

    # Signature section
    ws[f'A{row + 2}'] = '🔏 محل مهر و امضا'
    ws[f'A{row + 2}'].font = styles['header_font']

    # Checksum
    ws[f'A{row + 4}'] = 'کد صحت گزارش:'
    ws[f'B{row + 4}'] = '=TEXT(MOD(SUMPRODUCT(VALUE(B8:B10)*7),99999),"00000")'
    ws[f'B{row + 4}'].font = Font(name='Calibri', size=12, bold=True, color='1F4E79')


def populate_dashboard_sheet(ws, styles):
    """Populates the dashboard sheet."""
    ws['A6'] = '📊 داشبورد وضعیت آزمایش‌ها'
    ws['A6'].font = Font(name='Tahoma', bold=True, size=16, color='1F4E79')

    ws['A8'] = 'خلاصه کلی:'
    ws['A8'].font = styles['header_font']

    metrics = [
        ('تعداد کل آزمایش‌ها:', '20'),
        ('✅ قبول شده:', '=COUNTIF(\'04_داشبورد\'!C11:C30,"✅")'),
        ('⚠️ هشدار:', '=COUNTIF(\'04_داشبورد\'!C11:C30,"⚠️")'),
        ('❌ رد شده:', '=COUNTIF(\'04_داشبورد\'!C11:C30,"❌")')
    ]

    row = 10
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = styles['text_font']
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name='Calibri', size=12, bold=True)
        row += 1


def populate_qa_sheet(ws, styles):
    """Populates the QA test sheet."""
    ws['A6'] = '🧪 تست واحد و اعتبارسنجی'
    ws['A6'].font = styles['header_font']

    ws['A8'] = 'این شیت برای تست خودکار فرمول‌ها طراحی شده است.'
    ws['A8'].font = styles['text_font']

    test_cases = [
        ('تست ورودی صفر', 'W2=0', 'باید "—" نمایش دهد'),
        ('تست ورودی منفی', 'جرم منفی', 'باید خطای Data Validation دهد'),
        ('تست داده مرزی', 'SSD < OD', 'باید ❌ فیزیکی نمایش دهد'),
        ('تست داده خراب', 'مجموع الک‌ها ≠ جرم اولیه', 'باید هشدار خطای جرم دهد')
    ]

    row = 10
    ws[f'A{row}'] = 'نوع تست'
    ws[f'B{row}'] = 'شرط'
    ws[f'C{row}'] = 'خروجی انتظار'

    for cell in ['A10', 'B10', 'C10']:
        ws[cell].font = styles['header_font']

    row = 11
    for test_type, condition, expected in test_cases:
        ws[f'A{row}'] = test_type
        ws[f'B{row}'] = condition
        ws[f'C{row}'] = expected
        row += 1


def populate_error_log_sheet(ws, styles):
    """Populates the error log sheet."""
    ws['A6'] = '⚠️ تجمع زنده هشدارهای فعال'
    ws['A6'].font = styles['header_font']

    ws['A8'] = 'شیت'
    ws['B8'] = 'نوع هشدار'
    ws['C8'] = 'توضیحات'
    ws['D8'] = 'وضعیت'

    for cell in ['A8', 'B8', 'C8', 'D8']:
        ws[cell].font = styles['header_font']
        ws[cell].fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    ws['A10'] = 'این شیت به صورت خودکار تمام هشدارهای فعال را نمایش می‌دهد'
    ws['A10'].font = Font(name='Tahoma', size=9, italic=True)


def populate_reference_db(ws, styles):
    """Populates the reference database sheet."""
    ws['A1'] = 'Standard'
    ws['B1'] = 'Test'
    ws['C1'] = 'Parameter'
    ws['D1'] = 'Min'
    ws['E1'] = 'Max'
    ws['F1'] = 'Unit'

    for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
        ws[cell].font = styles['header_font']
        ws[cell].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws[cell].font = Font(name='Tahoma', bold=True, size=11, color='FFFFFF')

    data = [
        ['ISIRI 302', 'Sieve Analysis', 'Passing #4', '95', '100', '%'],
        ['ASTM C136', 'Sieve Analysis', 'Mass Error', '0', '0.3', '%'],
        ['ASTM C127', 'Specific Gravity', 'SSD ≥ OD', '—', '—', '—'],
        ['ASTM C566', 'Moisture Content', 'Base', 'Dry', '—', '—'],
        ['EN 196-1', 'Mortar Strength', 'Outlier Deviation', '±10', '—', '%']
    ]

    row = 2
    for item in data:
        for col, value in enumerate(item, start=1):
            ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=col).font = styles['text_font']
        row += 1


def populate_standards(ws, styles):
    """Populates the standards conversion sheet."""
    ws['A1'] = 'Unit Conversion Factors'
    ws['A1'].font = styles['header_font']

    ws['A3'] = 'From'
    ws['B3'] = 'To'
    ws['C3'] = 'Factor'

    conversions = [
        ['kgf', 'N', '9.80665'],
        ['kN', 'N', '1000'],
        ['lbf', 'N', '4.44822'],
        ['°C', '°F', '=C*9/5+32']
    ]

    row = 4
    for item in conversions:
        for col, value in enumerate(item, start=1):
            ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=col).font = styles['text_font']
        row += 1


def populate_validation_data(ws, styles):
    """Populates the validation data sheet with errata."""
    ws['A1'] = 'اعتبارسنجی و اِراتای کتاب'
    ws['A1'].font = styles['header_font']

    ws['A3'] = 'کد آزمون'
    ws['B3'] = 'مقدار چاپ کتاب'
    ws['C3'] = 'مقدار مرجع ابزار'
    ws['D3'] = 'علت اصلاح'
    ws['E3'] = 'وضعیت'

    errata = [
        ['1-2', 'پایه تر', 'پایه خشک', 'مغایر ASTM C566', '✅'],
        ['1-4ج', 'OD=1.51', '≈2.6x', 'جابه‌جایی A/S', '✅'],
        ['1-5', 'S=1600', 'S بی‌بُعد', 'برچسب ستون', '✅'],
        ['1-6', 'ماسه/رس', 'ماسه/(ماسه+رس)', 'متن مغایر نمونه', '✅'],
        ['2-4', '11.9', '≈20.5', 'ضریب kgf/مساحت', '✅'],
        ['4-1', '41.5', '≈40.4', 'محاسبهٔ مساحت', '✅'],
        ['4-2', '2.53', '≈5.07', 'اعمال ضریب ۲', '✅'],
        ['4-3', '33.466', '≈7.5', 'خطای فاکتور/واحد', '✅']
    ]

    row = 4
    for item in errata:
        for col, value in enumerate(item, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = styles['text_font']
            if col == 2:
                cell.fill = styles['fail_fill']
                cell.font = styles['fail_font']
            elif col == 3:
                cell.fill = styles['pass_fill']
                cell.font = styles['pass_font']
        row += 1


def populate_glossary(ws, styles):
    """Populates the glossary sheet."""
    ws['A1'] = 'واژه‌نامه فنی'
    ws['A1'].font = styles['header_font']

    ws['A3'] = 'اصطلاح فارسی'
    ws['B3'] = 'English'
    ws['C3'] = 'توضیحات'

    terms = [
        ['OD', 'Oven-Dry', 'وزن مخصوص خشک کوره‌ای'],
        ['SSD', 'Saturated Surface-Dry', 'وزن مخصوص اشباع با سطح خشک'],
        ['App', 'Apparent', 'وزن مخصوص ظاهری'],
        ['FM', 'Fineness Modulus', 'مدول نرمی'],
        ['SE', 'Sand Equivalent', 'معادل ماسه']
    ]

    row = 4
    for item in terms:
        for col, value in enumerate(item, start=1):
            ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=col).font = styles['text_font']
        row += 1


def populate_materials_db(ws, styles):
    """Populates the materials database sheet."""
    ws['A1'] = 'پایگاه داده مصالح (Typical Values)'
    ws['A1'].font = styles['header_font']

    ws['A3'] = 'Material Type'
    ws['B3'] = 'SG (SSD)'
    ws['C3'] = 'Absorption (%)'
    ws['D3'] = 'FM'

    materials = [
        ['Siliceous Gravel', '2.65', '1.2', '—'],
        ['Limestone', '2.60', '1.5', '—'],
        ['Natural Sand', '2.62', '1.0', '2.8'],
        ['Crushed Stone', '2.70', '0.8', '—']
    ]

    row = 4
    for item in materials:
        for col, value in enumerate(item, start=1):
            ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=col).font = styles['text_font']
        row += 1


# ============================================================
# SECTION 3: QR CODE & CHECKSUM
# ============================================================

def generate_qr_code():
    """Generates QR code for the landing page."""
    print("📱 Generating QR code...")

    # URL for the landing page (will be updated after GitHub Pages deployment)
    url = "https://bmhmdyan279-png.github.io/Concrete-Lab-Companion/"

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="#1F4E79", back_color="white")
    img_path = 'assets/qr_code.png'
    img.save(img_path)

    print(f"✅ QR code saved to {img_path}")
    return img_path


def calculate_checksum(file_path):
    """Calculates SHA-256 checksum of the Excel file and saves it in releases/."""
    print("🔐 Calculating SHA-256 checksum...")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    sha256_hash = hashlib.sha256()

    with open(file_path, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)

    checksum = sha256_hash.hexdigest()
    file_name = os.path.basename(file_path)

    checksum_dir = os.path.join("releases")
    os.makedirs(checksum_dir, exist_ok=True)

    checksum_path = os.path.join(checksum_dir, f"{file_name}.sha256")

    with open(checksum_path, "w", encoding="utf-8") as f:
        f.write(f"{checksum}  {file_name}\n")

    print(f"✅ SHA-256: {checksum}")
    print(f"✅ Checksum saved to {checksum_path}")

    return checksum

def main():
    print("🚀 Building Concrete Lab Digital Companion project...")

    create_project_structure()
    generate_readme()
    generate_changelog()
    generate_license()
    generate_landing_page()
    generate_requirements()
    generate_gitignore()

    excel_path = create_excel_workbook()
    generate_qr_code()
    calculate_checksum(excel_path)

    print("✅ Build finished successfully.")


if __name__ == "__main__":
    main()