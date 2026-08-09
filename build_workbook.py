import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os

# --- تنظیمات استایل (منطبق بر سند نهایی بند 2) ---
COLORS = {
    'input': 'FFF2CC',    # زرد ورودی
    'calc': 'F2F2F2',     # خاکستری محاسبه
    'pass': 'C6EFCE',     # سبز قبولی
    'warn': 'FCE4D6',     # نارنجی هشدار
    'fail': 'FFC7CE',     # قرمز خطا
    'text_pass': '006100',
    'text_fail': '9C0006'
}

fill_input = PatternFill(start_color=COLORS['input'], end_color=COLORS['input'], fill_type="solid")
fill_calc = PatternFill(start_color=COLORS['calc'], end_color=COLORS['calc'], fill_type="solid")
font_normal = Font(name='Tahoma', size=11)
font_bold = Font(name='Tahoma', size=11, bold=True)
align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

def create_sheet(wb, name, title, headers, is_hidden=False):
    ws = wb.create_sheet(name)
    ws.sheet_view.rightToLeft = True
    if is_hidden: ws.sheet_state = 'hidden'
    
    ws['A1'] = title
    ws['A1'].font = Font(name='Tahoma', size=14, bold=True, color='1F4E79')
    ws.merge_cells('A1:E1')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = font_bold
        cell.fill = fill_calc
        cell.border = Border(bottom=Side(style='thin'))
        cell.alignment = align_center
    return ws

# --- شروع ساخت ورک‌بوک ---
wb = openpyxl.Workbook()
wb.remove(wb.active)

# 1. شیت راهنما (00_راهنما)
ws_guide = create_sheet(wb, "00_راهنما", "همراه دیجیتال آزمایشگاه بتن (v1.0.0)", ["بخش", "توضیحات"])
guide_data = [
    ("فلسفه", "آموزش + محاسبه + کنترل کیفیت (QA)"),
    ("رمز عبور", "ConcreteLab2026! (فقط در مخزن گیت‌هاب)"),
    ("اِراتایاب", "شیت پنهان _Validation_Data حاوی ۸ خطای چاپ اول کتاب"),
    ("نمودار", "Plan A: Scatter لگاریتمی | Plan B: ستون Helper برای موبایل")
]
for r, (k, v) in enumerate(guide_data, 4):
    ws_guide.cell(r, 1, k).font = font_bold
    ws_guide.cell(r, 2, v)

# 2. شیت اطلاعات آزمون (01_اطلاعات_آزمون)
ws_info = create_sheet(wb, "01_اطلاعات_آزمون", "شناسنامه آزمایش", ["فیلد", "مقدار"])
info_fields = ["پروژه", "شماره نمونه", "تاریخ", "اپراتور", "دستگاه", "دما (C)"]
for r, f in enumerate(info_fields, 4):
    ws_info.cell(r, 1, f).font = font_bold
    c = ws_info.cell(r, 2)
    c.fill = fill_input
    c.protection = Protection(locked=False)

# 3. شیت‌های آزمایش (02_آزمایش‌ها) - نمونه‌های کلیدی
# شیت 1-2 رطوبت (اصلاح شده پایه خشک)
ws_1_2 = create_sheet(wb, "1-2 رطوبت", "درصد رطوبت (ASTM C566 - پایه خشک)", ["شرح", "وزن (g)", "فرمول/نتیجه"])
ws_1_2.cell(4, 1, "وزن تر (W1)").font = font_bold
ws_1_2.cell(4, 2).fill = fill_input; ws_1_2.cell(4, 2).protection = Protection(locked=False)
ws_1_2.cell(5, 1, "وزن خشک (W2)").font = font_bold
ws_1_2.cell(5, 2).fill = fill_input; ws_1_2.cell(5, 2).protection = Protection(locked=False)
ws_1_2.cell(6, 1, "رطوبت (%)").font = font_bold
ws_1_2.cell(6, 2).fill = fill_calc
ws_1_2['C6'] = '=IF(B5=0,"—",ROUND((B4-B5)/B5*100,2))' # فرمول اصلاح شده منتقد 3

# شیت 2-4 مقاومت ملات (با حذف داده پرت EN 196-1)
ws_2_4 = create_sheet(wb, "2-4 ملات", "مقاومت ملات (EN 196-1)", ["نمونه", "بار (kgf)", "MPa"])
for i in range(1, 7):
    ws_2_4.cell(3+i, 1, f"نمونه {i}").font = font_bold
    ws_2_4.cell(3+i, 2).fill = fill_input; ws_2_4.cell(3+i, 2).protection = Protection(locked=False)
    ws_2_4.cell(3+i, 3).fill = fill_calc
    # فرمول: تبدیل kgf به N و تقسیم بر 1600mm2
    ws_2_4.cell(3+i, 3).value = f'=IF(B{3+i}="","",ROUND(B{3+i}*9.80665/1600,1))'

# 4. لایه پنهان اعتبارسنجی (_Validation_Data)
ws_val = create_sheet(wb, "_Validation_Data", "اِراتایاب چاپ اول", ["کد", "مقدار کتاب", "مقدار ابزار", "وضعیت"], True)
errata = [
    ("1-2", "پایه تر", "پایه خشک", "اصلاح فرمول"),
    ("1-4ج", "1.51", "2.6x", "غیرممکن فیزیکی"),
    ("2-4", "11.9", "20.5", "خطای فاکتور kgf"),
    ("4-3", "33.466", "7.5", "خطای فرمول خمشی")
]
for r, (c, b, t, n) in enumerate(errata, 4):
    ws_val.cell(r, 1, c); ws_val.cell(r, 2, b); ws_val.cell(r, 3, t); ws_val.cell(r, 4, n)

# 5. شیت گزارش (03_گزارش)
ws_rep = create_sheet(wb, "03_گزارش", "گزارش نهایی آزمایشگاه", ["آزمایش", "نتیجه", "وضعیت"])
ws_rep.cell(4, 1, "رطوبت").font = font_bold
ws_rep['B4'] = "='1-2 رطوبت'!C6"
ws_rep['C4'] = '=IF(B4="","انجام نشده","✅")'

# --- قفل‌گذاری نهایی ورک‌بوک ---
for ws in wb.worksheets:
    if ws.sheet_state != 'hidden':
        ws.protection.sheet = True
        ws.protection.password = 'ConcreteLab2026!'
        ws.protection.enable()

# ذخیره فایل
os.makedirs('releases', exist_ok=True)
file_name = 'releases/Concrete_Lab_Digital_Companion_v1.0.0.xlsx'
wb.save(file_name)
print(f"✅ ورک‌بوک نهایی با موفقیت در {file_name} تولید شد.")
